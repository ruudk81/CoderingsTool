# === MODULES ========================================================================================================
import os
from typing import Any, List, Dict, Optional
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# === MODELS ========================================================================================================
import models
from development.step_6_codeAssigner.models_codeAssigner import CodeAssignedModel
from development.step_5_codeGenerator.prompts_codeGenerator import ConsolidatedCode
from development.step_4_classifier.models_classifier import DomainSet, DomainResultModel

# === UTILS ========================================================================================================
from utils.verboseReporter import VerboseReporter


# === CANONICAL COLUMN DEFINITIONS (single source of truth) ========================================================
# (dict_key, excel_header, column_width)
EXPORT_COLUMNS = [
    ('respondent_id',         'Respondent ID',         15),
    ('original_response',     'Original Response',     50),
    ('idea_id',               'Idea ID',               15),
    ('idea_text',             'Idea Text',             50),
    ('instance',              'Verbatim Instance',     40),
    ('domain',                'Domain',                20),
    ('facet',                 'Facet',                 20),
    ('attribute',             'Attribute',             20),
    ('valence',               'Valence',               10),
    ('code_label',            'Code Label',            30),
    ('code_description',      'Code Description',      50),
    ('theme_name',            'Theme Name',            30),
    ('theme_description',     'Theme Description',     50),
    ('category',              'Category (Facet)',      30),
    ('category_description',  'Category Description',  50),
    ('assignment_confidence', 'Assignment Confidence', 20),
    ('assignment_rationale',  'Assignment Rationale',  60),
]


class ResultsExporter:

    def __init__(self, verbose: bool = True):
        self.verbose = verbose
        self.verbose_reporter = VerboseReporter(verbose, capture_logging=True)

    def _sanitize_excel_text(self, text):
        """Remove or replace control characters that Excel cannot handle"""
        if text is None or not isinstance(text, str):
            return text
        sanitized = ''.join(char if ord(char) > 31 or char in '\t\n\r' else ' ' for char in text)
        sanitized = sanitized.replace('\x08', '')
        return sanitized

    def _empty_row(self, columns):
        """Create a row dict with all keys from the given column spec set to empty string."""
        return {key: '' for key, _, _ in columns}

    def _build_code_to_info(self, codes: List[ConsolidatedCode]) -> Dict[str, Dict[str, Any]]:
        """Build code-to-info mapping from ConsolidatedCode list."""
        code_to_info = {}
        for code in codes:
            code_to_info[code.code_name] = {
                'code_description': code.definition,
                'valence': code.valence,
                'diagnostic_test': code.diagnostic_test,
            }
        return code_to_info

    def _build_partition_lookup(self, partition_set: DomainSet) -> Dict[str, str]:
        """Build partition_name → inclusion_definition lookup."""
        return {
            p.partition_name: p.inclusion_definition
            for p in partition_set.partitions
        }

    def _build_facet_desc_lookup(self, partition_results: Dict[str, DomainResultModel]) -> Dict[str, str]:
        """Build facet_name → facet_description lookup from all partitions."""
        lookup = {}
        for domain_name, result in partition_results.items():
            for facet_dict in result.facets:
                fname = facet_dict.get('facet_name', '')
                fdesc = facet_dict.get('facet_description', '')
                if fname:
                    lookup[fname] = fdesc
        return lookup

    def export_to_excel(self,
                       code_assigned_results: List[CodeAssignedModel],
                       codes: List[ConsolidatedCode],
                       partition_set: DomainSet,
                       partition_results: Dict[str, DomainResultModel],
                       filename: str,
                       var_name: str,
                       quality_filtered_text: Optional[List] = None,
                       export_dir: Optional[str] = None) -> str:
        """Export code assignment results to Excel.

        Columns: respondent_id, original_response, idea_id, idea_text,
        instance, domain, facet, attribute, valence, code_label, code_description,
        theme_name, theme_description, category, category_description,
        assignment_confidence, assignment_rationale.
        """

        self.verbose_reporter.section_header("EXPORTING CODE ASSIGNMENTS TO EXCEL")

        export_data = []
        code_to_info = self._build_code_to_info(codes)
        partition_lookup = self._build_partition_lookup(partition_set)
        facet_desc_lookup = self._build_facet_desc_lookup(partition_results)

        # Process each result
        for result in code_assigned_results:
            respondent_id = result.respondent_id

            if result.response_ideas:
                for idea in result.response_ideas:
                    code_info = code_to_info.get(idea.assigned_code) if idea.assigned_code else None

                    # Theme info from partition
                    theme_name = idea.partition_name or ''
                    theme_description = partition_lookup.get(theme_name, '')

                    # Category info from facet
                    category = idea.facet or ''
                    category_description = facet_desc_lookup.get(category, '')

                    row = self._empty_row(EXPORT_COLUMNS)
                    row.update({
                        'respondent_id': respondent_id,
                        'original_response': result.response,
                        'idea_id': idea.idea_id,
                        'idea_text': idea.idea,
                        'instance': getattr(idea, 'instance', ''),
                        'domain': getattr(idea, 'domain', ''),
                        'facet': category,
                        'attribute': getattr(idea, 'attribute', ''),
                        'valence': getattr(idea, 'valence', ''),
                        'code_label': idea.assigned_code or 'No Code Assigned',
                        'code_description': code_info['code_description'] if code_info else '',
                        'theme_name': theme_name,
                        'theme_description': theme_description,
                        'category': category,
                        'category_description': category_description,
                        'assignment_confidence': idea.confidence,
                        'assignment_rationale': idea.rationale or '',
                    })
                    export_data.append(row)

        # Add filtered responses if provided
        if quality_filtered_text:
            FILTER_CODE_LABELS = {
                99999997: "User Missing (Don't Know)",
                99999998: "System Missing (NA)",
                99999999: "No Answer (Meaningless)"
            }
            filtered_responses = [r for r in quality_filtered_text if r.quality_filter]
            for response in filtered_responses:
                filter_code = response.quality_filter_code
                filter_label = FILTER_CODE_LABELS.get(filter_code, f"Unknown Filter ({filter_code})")
                row = self._empty_row(EXPORT_COLUMNS)
                row.update({
                    'respondent_id': response.respondent_id,
                    'original_response': response.response,
                    'code_label': str(filter_code),
                    'code_description': filter_label,
                    'theme_name': 'FILTERED',
                    'theme_description': filter_label,
                    'assignment_confidence': 1.0,
                    'assignment_rationale': 'Filtered in quality check',
                })
                export_data.append(row)

        # Convert to DataFrame with canonical column order
        df = pd.DataFrame(export_data, columns=[key for key, _, _ in EXPORT_COLUMNS])

        # Create export directory
        if export_dir is None:
            project_root = Path(__file__).parent.parent.parent.parent
            export_dir = os.path.join(project_root, 'exports')
        Path(export_dir).mkdir(parents=True, exist_ok=True)

        # Create output filename
        base_name = Path(filename).stem
        output_filename = f"{base_name}_{var_name}_code_assignments.xlsx"
        output_path = os.path.join(export_dir, output_filename)

        print(f"About to export DataFrame with shape: {df.shape}")

        # Export to Excel with formatting
        self._write_formatted_excel(df, output_path, var_name, EXPORT_COLUMNS)

        # Report statistics
        self.verbose_reporter.stat_line(f"Total rows exported: {len(export_data)}")
        self.verbose_reporter.stat_line(f"Unique respondents: {df['respondent_id'].nunique()}")
        self.verbose_reporter.stat_line(f"Unique ideas: {df['idea_id'].nunique()}")
        self.verbose_reporter.stat_line(f"Unique codes assigned: {df[df['code_label'] != 'No Code Assigned']['code_label'].nunique()}")

        return output_path

    def _write_formatted_excel(self, df: pd.DataFrame, output_path: str, var_name: str, columns: list):
        """Write DataFrame to Excel with formatting. Column spec drives headers, widths, and confidence formatting."""

        wb = Workbook()
        ws = wb.active
        ws.title = "codering"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers from canonical column spec
        headers = [header for _, header, _ in columns]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Find confidence column index dynamically
        confidence_col_idx = next(
            (i for i, (key, _, _) in enumerate(columns, 1) if key == 'assignment_confidence'),
            None
        )

        # Write data
        try:
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
                for c_idx, value in enumerate(row, 1):
                    if value is None:
                        value = ""
                    if isinstance(value, str):
                        value = self._sanitize_excel_text(value)
                    try:
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = border
                    except Exception as cell_error:
                        print(f"DEBUG: Error writing cell at row {r_idx}, col {c_idx}: {cell_error}")
                        raise

                    # Format confidence as float
                    if c_idx == confidence_col_idx and value is not None and value != '':
                        try:
                            cell.value = float(value)
                            cell.number_format = '0.00'
                        except (ValueError, TypeError):
                            pass
        except Exception as e:
            print(f"DEBUG: Error during data writing: {e}")
            raise

        # Column widths from spec
        for i, (_, _, width) in enumerate(columns):
            col_letter = chr(ord('A') + i) if i < 26 else chr(ord('A') + i // 26 - 1) + chr(ord('A') + i % 26)
            ws.column_dimensions[col_letter].width = width

        ws.freeze_panes = 'A2'

        # Summary sheet
        summary_ws = wb.create_sheet(title="Summary")
        summary_data = [
            ["Summary Statistics", ""],
            ["", ""],
            ["Total Assignments", len(df)],
            ["Unique Respondents", df['respondent_id'].nunique()],
            ["Unique Ideas", df['idea_id'].nunique()],
            ["Unique Codes", df[df['code_label'] != 'No Code Assigned']['code_label'].nunique()],
            ["Unique Themes", df[df['theme_name'] != '']['theme_name'].nunique()],
            ["", ""],
            ["Code Frequency", "Count"],
        ]

        code_freq = df[df['code_label'] != 'No Code Assigned']['code_label'].value_counts()
        for code, count in code_freq.items():
            summary_data.append([code, count])

        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1 or (row_idx == 9 and col_idx == 1):
                    cell.font = Font(bold=True, size=14)
                elif row_idx in [3, 4, 5, 6, 7]:
                    if col_idx == 1:
                        cell.font = Font(bold=True)

        summary_ws.column_dimensions['A'].width = 30
        summary_ws.column_dimensions['B'].width = 15

        try:
            wb.save(output_path)
        except Exception as e:
            print(f"DEBUG: Error saving workbook: {e}")
            raise
