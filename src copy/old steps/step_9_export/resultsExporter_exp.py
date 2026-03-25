import os, sys; sys.path.extend([p for p in [os.getcwd().split('coderingsTool')[0] + suffix for suffix in ['', 'coderingsTool', 'coderingsTool/src', 'coderingsTool/src/utils']] if p not in sys.path]) if 'coderingsTool' in os.getcwd() else None

# === MODULES ========================================================================================================
import os
from typing import Any, List, Dict, Optional
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# === MODELS ========================================================================================================
import models
from development.models_exp import CodeAssignedModel as ExpCodeAssignedModel

# === UTILS ========================================================================================================
from utils.verboseReporter import VerboseReporter
from utils.codeGenerator import CodeGeneratorReasoningResults


# === CANONICAL COLUMN DEFINITIONS (single source of truth) ========================================================
# (dict_key, excel_header, column_width)
EXPORT_COLUMNS = [
    ('respondent_id',         'Respondent ID',         15),
    ('original_response',     'Original Response',     50),
    ('idea_id',               'Idea ID',               15),
    ('idea_text',             'Idea Text',             50),
    ('instance',              'Verbatim Instance',     40),
    ('concept_type',          'Concept Type',          20),
    ('valence',               'Valence',               10),
    ('code_label',            'Code Label',            30),
    ('code_description',      'Code Description',      50),
    ('theme_name',            'Theme Name',            30),
    ('theme_description',     'Theme Description',     50),
    ('category',              'Category',              30),
    ('category_description',  'Category Description',  50),
    ('assignment_confidence', 'Assignment Confidence', 20),
    ('assignment_rationale',  'Assignment Rationale',  60),
]

EXPORT_COLUMNS_WITH_REASONING = EXPORT_COLUMNS + [
    ('codegen_theme',          'Codegen Theme',          60),
    ('codegen_recommendation', 'Codegen Recommendation', 60),
    ('codebook_validation',    'Codebook Validation',    60),
]


class ResultsExporter:

    def __init__(self, verbose: bool = True):
        self.verbose = verbose
        self.verbose_reporter = VerboseReporter(verbose, capture_logging=True)

    def _sanitize_excel_text(self, text):
        """Remove or replace control characters that Excel cannot handle"""
        if text is None or not isinstance(text, str):
            return text

        # Remove all control characters (0x00-0x1F) except tab, newline, carriage return
        sanitized = ''.join(char if ord(char) > 31 or char in '\t\n\r' else ' ' for char in text)
        sanitized = sanitized.replace('\x08', '')

        return sanitized

    def _empty_row(self, columns):
        """Create a row dict with all keys from the given column spec set to empty string."""
        return {key: '' for key, _, _ in columns}

    def _build_code_to_info(self, theme_enriched_codebook):
        """Build code-to-info mapping from the enriched codebook."""
        code_to_info = {}
        for code_entry in theme_enriched_codebook.codes:
            code_to_info[code_entry.code] = {
                'theme_name': code_entry.theme or 'No Theme',
                'theme_description': code_entry.theme_description or 'No Description',
                'category': code_entry.category or '',
                'category_description': getattr(code_entry, 'category_description', '') or '',
                'code_description': code_entry.definition,
                'source_cluster': code_entry.source_cluster,
            }
        return code_to_info

    def export_to_excel(self,
                       code_assigned_results: List[ExpCodeAssignedModel],
                       theme_enriched_codebook: models.ThemeEnrichedCodebookModel,
                       filename: str,
                       var_name: str,
                       quality_filtered_text: Optional[List] = None,
                       export_dir: Optional[str] = None,
                       include_visualizations: bool = False,
                       clustering_metadata: Optional[Any] = None,
                       extraction_metadata: Optional[Any] = None) -> str:
        """Export code assignment results to Excel.

        Columns: respondent_id, original_response, idea_id, idea_text,
        instance, concept_type, valence, code_label, code_description,
        theme_name, theme_description, category, category_description,
        assignment_confidence, assignment_rationale.
        """

        self.verbose_reporter.section_header("EXPORTING CODE ASSIGNMENTS TO EXCEL")

        export_data = []
        code_to_info = self._build_code_to_info(theme_enriched_codebook)

        # Process each result
        for result in code_assigned_results:
            respondent_id = result.respondent_id

            if result.response_ideas:
                for idea in result.response_ideas:
                    if idea.assigned_codes:
                        for idx, code in enumerate(idea.assigned_codes):
                            code_info = code_to_info.get(code)

                            # Fallback: use assigned_themes from step 8 if code not in codebook
                            if code_info is None:
                                fallback_theme = ''
                                if idea.assigned_themes and idx < len(idea.assigned_themes):
                                    fallback_theme = idea.assigned_themes[idx]
                                code_info = {
                                    'theme_name': fallback_theme or 'Unknown Theme',
                                    'theme_description': '',
                                    'category': '',
                                    'category_description': '',
                                    'code_description': 'Unknown Code',
                                }

                            row = self._empty_row(EXPORT_COLUMNS)
                            row.update({
                                'respondent_id': respondent_id,
                                'original_response': result.response,
                                'idea_id': idea.idea_id,
                                'idea_text': idea.idea,
                                'instance': getattr(idea, 'instance', ''),
                                'concept_type': getattr(idea, 'concept_type', ''),
                                'valence': getattr(idea, 'valence', ''),
                                'code_label': code,
                                'code_description': code_info['code_description'],
                                'theme_name': code_info['theme_name'],
                                'theme_description': code_info['theme_description'],
                                'category': code_info['category'],
                                'category_description': code_info['category_description'],
                                'assignment_confidence': getattr(idea, 'assignment_confidence', None),
                                'assignment_rationale': getattr(idea, 'assignment_rationale', ''),
                            })
                            export_data.append(row)
                    else:
                        # No codes assigned
                        row = self._empty_row(EXPORT_COLUMNS)
                        row.update({
                            'respondent_id': respondent_id,
                            'original_response': result.response,
                            'idea_id': idea.idea_id,
                            'idea_text': idea.idea,
                            'instance': getattr(idea, 'instance', ''),
                            'concept_type': getattr(idea, 'concept_type', ''),
                            'valence': getattr(idea, 'valence', ''),
                            'code_label': 'No Code Assigned',
                            'assignment_confidence': getattr(idea, 'assignment_confidence', None),
                            'assignment_rationale': getattr(idea, 'assignment_rationale', ''),
                        })
                        export_data.append(row)

        # Add filtered responses if provided
        if quality_filtered_text:
            FILTER_CODE_LABELS = {
                99999997: "User Missing (Don't Know)",
                99999998: "System Missing (NA)",
                99999999: "No Answer (Meaningless)"
            }
            filtered_responses = [r for r in quality_filtered_text if r.quality_filter]
            for response in filtered_responses:
                filter_code = response.quality_filter_code
                filter_label = FILTER_CODE_LABELS.get(filter_code, f"Unknown Filter ({filter_code})")
                row = self._empty_row(EXPORT_COLUMNS)
                row.update({
                    'respondent_id': response.respondent_id,
                    'original_response': response.response,
                    'code_label': str(filter_code),
                    'code_description': filter_label,
                    'theme_name': 'FILTERED',
                    'theme_description': filter_label,
                    'assignment_confidence': 1.0,
                    'assignment_rationale': 'Filtered in quality check',
                })
                export_data.append(row)

        # Convert to DataFrame with canonical column order
        df = pd.DataFrame(export_data, columns=[key for key, _, _ in EXPORT_COLUMNS])

        # Create export directory
        if export_dir is None:
            project_root = Path(__file__).parent.parent.parent.parent
            export_dir = os.path.join(project_root, 'exports')
        Path(export_dir).mkdir(parents=True, exist_ok=True)

        # Create output filename
        base_name = Path(filename).stem
        output_filename = f"{base_name}_{var_name}_code_assignments.xlsx"
        output_path = os.path.join(export_dir, output_filename)

        print(f"About to export DataFrame with shape: {df.shape}")

        # Export to Excel with formatting
        self._write_formatted_excel(df, output_path, var_name, EXPORT_COLUMNS)

        # Report statistics
        self.verbose_reporter.stat_line(f"Total rows exported: {len(export_data)}")
        self.verbose_reporter.stat_line(f"Unique respondents: {df['respondent_id'].nunique()}")
        self.verbose_reporter.stat_line(f"Unique ideas: {df['idea_id'].nunique()}")
        self.verbose_reporter.stat_line(f"Unique codes assigned: {df[df['code_label'] != 'No Code Assigned']['code_label'].nunique()}")

        # Add visualizations if requested
        if include_visualizations:
            try:
                from utils.exportVisualizer import ExportVisualizer
                from openpyxl import load_workbook

                self.verbose_reporter.stat_line("Generating visualizations...")

                visualizer = ExportVisualizer(
                    clustering_metadata=clustering_metadata,
                    code_assigned_results=code_assigned_results,
                    theme_enriched_codebook=theme_enriched_codebook,
                    extraction_metadata=extraction_metadata,
                    verbose=self.verbose
                )

                wb = load_workbook(output_path)
                wb = visualizer.add_visualizations_to_workbook(wb)
                wb.save(output_path)

                network_html_path = visualizer.generate_network_html(export_dir)
                if network_html_path:
                    self.verbose_reporter.stat_line(f"Network graph HTML: {network_html_path}")

            except ImportError as e:
                print(f"Warning: Could not import visualization dependencies: {e}")
                print("Install with: pip install wordcloud networkx")
            except Exception as e:
                print(f"Warning: Visualization generation failed: {e}")

        return output_path

    def export_to_excel_with_reasoning(self,
                                      code_assigned_results: List[ExpCodeAssignedModel],
                                      theme_enriched_codebook: models.ThemeEnrichedCodebookModel,
                                      reasoning_results: CodeGeneratorReasoningResults,
                                      filename: str,
                                      var_name: str,
                                      quality_filtered_text: Optional[List] = None,
                                      export_dir: Optional[str] = None) -> str:
        """Export code assignment results with step 7 reasoning data to Excel."""

        self.verbose_reporter.section_header("EXPORTING CODE ASSIGNMENTS WITH REASONING TO EXCEL")

        export_data = []
        code_to_info = self._build_code_to_info(theme_enriched_codebook)
        reasoning_mapping = self._create_reasoning_mapping(reasoning_results)

        columns = EXPORT_COLUMNS_WITH_REASONING

        for result in code_assigned_results:
            respondent_id = result.respondent_id

            if result.response_ideas:
                for idea in result.response_ideas:
                    if idea.assigned_codes:
                        for idx, code in enumerate(idea.assigned_codes):
                            code_info = code_to_info.get(code)

                            if code_info is None:
                                fallback_theme = ''
                                if idea.assigned_themes and idx < len(idea.assigned_themes):
                                    fallback_theme = idea.assigned_themes[idx]
                                code_info = {
                                    'theme_name': fallback_theme or 'Unknown Theme',
                                    'theme_description': '',
                                    'category': '',
                                    'category_description': '',
                                    'code_description': 'Unknown Code',
                                    'source_cluster': None,
                                }

                            source_cluster = code_info.get('source_cluster')
                            parent_cluster = source_cluster.split('-')[0] if source_cluster and isinstance(source_cluster, str) and '-' in source_cluster else source_cluster
                            reasoning_data = self._get_reasoning_for_cluster(reasoning_mapping, source_cluster, parent_cluster)

                            row = self._empty_row(columns)
                            row.update({
                                'respondent_id': respondent_id,
                                'original_response': result.response,
                                'idea_id': idea.idea_id,
                                'idea_text': idea.idea,
                                'instance': getattr(idea, 'instance', ''),
                                'concept_type': getattr(idea, 'concept_type', ''),
                                'valence': getattr(idea, 'valence', ''),
                                'code_label': code,
                                'code_description': code_info['code_description'],
                                'theme_name': code_info['theme_name'],
                                'theme_description': code_info['theme_description'],
                                'category': code_info['category'],
                                'category_description': code_info['category_description'],
                                'assignment_confidence': getattr(idea, 'assignment_confidence', None),
                                'assignment_rationale': getattr(idea, 'assignment_rationale', ''),
                                'codegen_theme': reasoning_data.get('codegen_theme', ''),
                                'codegen_recommendation': reasoning_data.get('codegen_recommendation', ''),
                                'codebook_validation': reasoning_data.get('codebook_validation', ''),
                            })
                            export_data.append(row)
                    else:
                        initial_cluster = getattr(idea, 'initial_cluster', None)
                        parent_cluster = initial_cluster.split('-')[0] if initial_cluster and isinstance(initial_cluster, str) and '-' in initial_cluster else initial_cluster
                        reasoning_data = self._get_reasoning_for_cluster(reasoning_mapping, initial_cluster, parent_cluster)

                        row = self._empty_row(columns)
                        row.update({
                            'respondent_id': respondent_id,
                            'original_response': result.response,
                            'idea_id': idea.idea_id,
                            'idea_text': idea.idea,
                            'instance': getattr(idea, 'instance', ''),
                            'concept_type': getattr(idea, 'concept_type', ''),
                            'valence': getattr(idea, 'valence', ''),
                            'code_label': 'No Code Assigned',
                            'assignment_confidence': getattr(idea, 'assignment_confidence', None),
                            'assignment_rationale': getattr(idea, 'assignment_rationale', ''),
                            'codegen_theme': reasoning_data.get('codegen_theme', ''),
                            'codegen_recommendation': reasoning_data.get('codegen_recommendation', ''),
                            'codebook_validation': reasoning_data.get('codebook_validation', ''),
                        })
                        export_data.append(row)

        # Add filtered responses
        if quality_filtered_text:
            FILTER_CODE_LABELS = {
                99999997: "User Missing (Don't Know)",
                99999998: "System Missing (NA)",
                99999999: "No Answer (Meaningless)"
            }
            filtered_responses = [r for r in quality_filtered_text if r.quality_filter]
            for response in filtered_responses:
                filter_code = response.quality_filter_code
                filter_label = FILTER_CODE_LABELS.get(filter_code, f"Unknown Filter ({filter_code})")
                row = self._empty_row(columns)
                row.update({
                    'respondent_id': response.respondent_id,
                    'original_response': response.response,
                    'code_label': str(filter_code),
                    'code_description': filter_label,
                    'theme_name': 'FILTERED',
                    'theme_description': filter_label,
                    'assignment_confidence': 1.0,
                    'assignment_rationale': 'Filtered in quality check',
                })
                export_data.append(row)

        # Convert to DataFrame with canonical column order
        df = pd.DataFrame(export_data, columns=[key for key, _, _ in columns])

        # Create export directory
        if export_dir is None:
            project_root = Path(__file__).parent.parent.parent.parent
            export_dir = os.path.join(project_root, 'exports')
        Path(export_dir).mkdir(parents=True, exist_ok=True)

        base_name = Path(filename).stem
        output_filename = f"{base_name}_{var_name}_code_assignments_with_reasoning.xlsx"
        output_path = os.path.join(export_dir, output_filename)

        self._write_formatted_excel(df, output_path, var_name, columns)

        self.verbose_reporter.stat_line(f"Total rows exported: {len(export_data)}")
        self.verbose_reporter.stat_line(f"Unique respondents: {df['respondent_id'].nunique()}")
        self.verbose_reporter.stat_line(f"Unique ideas: {df['idea_id'].nunique()}")
        self.verbose_reporter.stat_line(f"Unique codes assigned: {df[df['code_label'] != 'No Code Assigned']['code_label'].nunique()}")
        self.verbose_reporter.stat_line(f"Excel file with reasoning saved: {output_path}")

        return output_path

    def _write_formatted_excel(self, df: pd.DataFrame, output_path: str, var_name: str, columns: list):
        """Write DataFrame to Excel with formatting. Column spec drives headers, widths, and confidence formatting."""

        wb = Workbook()
        ws = wb.active
        ws.title = "codering"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers from canonical column spec
        headers = [header for _, header, _ in columns]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Find confidence column index dynamically
        confidence_col_idx = next(
            (i for i, (key, _, _) in enumerate(columns, 1) if key == 'assignment_confidence'),
            None
        )

        # Write data
        try:
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
                for c_idx, value in enumerate(row, 1):
                    if value is None:
                        value = ""
                    if isinstance(value, str):
                        value = self._sanitize_excel_text(value)
                    try:
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = border
                    except Exception as cell_error:
                        print(f"DEBUG: Error writing cell at row {r_idx}, col {c_idx}: {cell_error}")
                        raise

                    # Format confidence as float
                    if c_idx == confidence_col_idx and value is not None and value != '':
                        try:
                            cell.value = float(value)
                            cell.number_format = '0.00'
                        except (ValueError, TypeError):
                            pass
        except Exception as e:
            print(f"DEBUG: Error during data writing: {e}")
            raise

        # Column widths from spec
        for i, (_, _, width) in enumerate(columns):
            col_letter = chr(ord('A') + i) if i < 26 else chr(ord('A') + i // 26 - 1) + chr(ord('A') + i % 26)
            ws.column_dimensions[col_letter].width = width

        ws.freeze_panes = 'A2'

        # Summary sheet
        summary_ws = wb.create_sheet(title="Summary")
        summary_data = [
            ["Summary Statistics", ""],
            ["", ""],
            ["Total Assignments", len(df)],
            ["Unique Respondents", df['respondent_id'].nunique()],
            ["Unique Ideas", df['idea_id'].nunique()],
            ["Unique Codes", df[df['code_label'] != 'No Code Assigned']['code_label'].nunique()],
            ["Unique Themes", df[df['theme_name'] != '']['theme_name'].nunique()],
            ["", ""],
            ["Code Frequency", "Count"],
        ]

        code_freq = df[df['code_label'] != 'No Code Assigned']['code_label'].value_counts()
        for code, count in code_freq.items():
            summary_data.append([code, count])

        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1 or (row_idx == 9 and col_idx == 1):
                    cell.font = Font(bold=True, size=14)
                elif row_idx in [3, 4, 5, 6, 7]:
                    if col_idx == 1:
                        cell.font = Font(bold=True)

        summary_ws.column_dimensions['A'].width = 30
        summary_ws.column_dimensions['B'].width = 15

        try:
            wb.save(output_path)
        except Exception as e:
            print(f"DEBUG: Error saving workbook: {e}")
            raise

    def _create_reasoning_mapping(self, reasoning_results: CodeGeneratorReasoningResults) -> Dict[str, Dict[str, str]]:
        """Create a mapping from cluster IDs to reasoning data"""
        reasoning_mapping = {}

        for cluster_id, step1_data in reasoning_results.step1_summaries.items():
            cluster_key = str(cluster_id)
            if cluster_key not in reasoning_mapping:
                reasoning_mapping[cluster_key] = {}
            analysis = step1_data.get('analysis', '')
            reasoning_mapping[cluster_key]['codegen_theme'] = analysis

        for cluster_id, step2_data in reasoning_results.step2_analysis.items():
            cluster_key = str(cluster_id)
            if cluster_key not in reasoning_mapping:
                reasoning_mapping[cluster_key] = {}
            if 'coding_decision' in step2_data and step2_data['coding_decision']:
                first_decision = step2_data['coding_decision']
                decision = first_decision.get('decision', '')
                justification = first_decision.get('justification', '')
                combined = f"{decision}: {justification}" if decision and justification else decision or justification
                reasoning_mapping[cluster_key]['codegen_recommendation'] = combined

        for cluster_id, step3_data in reasoning_results.step3_recommendations.items():
            cluster_key = str(cluster_id)
            if cluster_key not in reasoning_mapping:
                reasoning_mapping[cluster_key] = {}
            if 'generated_code' in step3_data and step3_data['generated_code']:
                first_code = step3_data['generated_code']
                code_label = first_code.get('code_label', '')
                code_definition = first_code.get('code_definition', '')
                combined = f"{code_label}: {code_definition}" if code_label and code_definition else code_label or code_definition
                reasoning_mapping[cluster_key]['generated_code'] = combined

        for cluster_id, step4_data in reasoning_results.step4_validations.items():
            cluster_key = str(cluster_id)
            if cluster_key not in reasoning_mapping:
                reasoning_mapping[cluster_key] = {}
            if 'code_validation' in step4_data and step4_data['code_validation']:
                first_validation = step4_data['code_validation']
                code_label = first_validation.get('code_label', '')
                justification = first_validation.get('decision_rationale', '')
                combined = f"{code_label}: {justification}" if code_label and justification else code_label or justification
                reasoning_mapping[cluster_key]['codebook_validation'] = combined

        return reasoning_mapping

    def _get_reasoning_for_cluster(self, reasoning_mapping: Dict[str, Dict[str, str]],
                                 source_cluster: Optional[str],
                                 parent_cluster: Optional[str]) -> Dict[str, str]:
        """Get reasoning data for a specific cluster, trying source cluster first then parent cluster"""
        default_reasoning = {'codegen_theme': '', 'codegen_recommendation': '', 'codebook_validation': ''}

        if source_cluster and str(source_cluster) in reasoning_mapping:
            return {**default_reasoning, **reasoning_mapping[str(source_cluster)]}
        if parent_cluster and str(parent_cluster) in reasoning_mapping:
            return {**default_reasoning, **reasoning_mapping[str(parent_cluster)]}

        return default_reasoning
