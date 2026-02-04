"""
ExportVisualizer - Generate visualizations for Excel export

Generates:
1. Dendrogram sheet (PNG in Excel) - Hierarchical tree of Theme > Category > Code > Cluster
2. Word Cloud grid (PNG in Excel) - Grid of word clouds for all codes based on taxonomy phrases
3. Network Graph (Interactive HTML) - Theme-Code-Cluster relationship network

Author: CoderingsTool
"""

import os
import sys
sys.path.extend([p for p in [os.getcwd().split('coderingsTool')[0] + suffix for suffix in ['', 'coderingsTool', 'coderingsTool/src', 'coderingsTool/src/utils']] if p not in sys.path]) if 'coderingsTool' in os.getcwd() else None

from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any, Union
from collections import defaultdict
import tempfile
import io

import numpy as np
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend for server/headless use
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.colors import to_hex

# Optional imports with graceful fallback
try:
    from wordcloud import WordCloud
    WORDCLOUD_AVAILABLE = True
except ImportError:
    WORDCLOUD_AVAILABLE = False
    print("Warning: wordcloud not installed. Word cloud generation will be skipped.")

try:
    import networkx as nx
    NETWORKX_AVAILABLE = True
except ImportError:
    NETWORKX_AVAILABLE = False
    print("Warning: networkx not installed. Network graph generation will be skipped.")

try:
    import plotly.graph_objects as go
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False
    print("Warning: plotly not installed. Interactive network graph will be skipped.")

try:
    import spacy
    SPACY_AVAILABLE = True
except ImportError:
    SPACY_AVAILABLE = False
    print("Warning: spacy not installed. Lemmatization will be skipped.")

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

import models
from .verboseReporter import VerboseReporter


@dataclass
class ExportVisualizerConfig:
    """Configuration for export visualizations."""

    # Dendrogram settings
    dendrogram_figsize: Tuple[int, int] = (24, 16)
    dendrogram_dpi: int = 150
    dendrogram_max_depth: int = 5  # Max hierarchy depth to show
    dendrogram_max_items_per_level: int = 50  # Truncate if too many items

    # Word cloud settings
    wordcloud_grid_cols: int = 4  # Columns in the grid
    wordcloud_cell_width: int = 400  # Width per word cloud
    wordcloud_cell_height: int = 300  # Height per word cloud
    wordcloud_max_words: int = 50  # Max words per cloud
    wordcloud_background: str = 'white'
    wordcloud_colormap: str = 'viridis'
    wordcloud_dpi: int = 100

    # Network graph settings
    network_figsize: Tuple[int, int] = (1200, 900)  # Plotly size in pixels
    network_layout: str = 'spring'  # 'spring', 'kamada_kawai', 'circular'
    network_theme_size: int = 40  # Node size for themes
    network_code_size: int = 25  # Node size for codes
    network_cluster_size: int = 15  # Node size for clusters

    # Output settings
    temp_dir: Optional[str] = None  # Directory for temporary files

    # Lemmatization settings
    spacy_model: str = 'nl_core_news_lg'  # or 'en_core_web_lg'
    pos_filter: str = 'all'  # 'all', 'nouns', 'adj_noun', 'bigrams'

    # Color palette for themes (will cycle if more themes)
    theme_colors: List[str] = field(default_factory=lambda: [
        '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
        '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf',
        '#aec7e8', '#ffbb78', '#98df8a', '#ff9896', '#c5b0d5'
    ])


class ExportVisualizer:
    """Generate visualizations for Excel export."""

    def __init__(
        self,
        clustering_metadata: Optional[models.ClusteringMetadataModel],
        code_assigned_results: List[models.CodeAssignedModel],
        theme_enriched_codebook: models.ThemeEnrichedCodebookModel,
        extraction_metadata: Optional[models.ExtractionMetadata] = None,
        config: Optional[ExportVisualizerConfig] = None,
        verbose: bool = True
    ):
        """
        Initialize visualizer with all required data sources.

        Args:
            clustering_metadata: Cluster-level data (keywords, labels, distributions)
            code_assigned_results: Ideas with assigned codes and taxonomy phrases
            theme_enriched_codebook: Hierarchical codebook (Theme > Category > Code)
            extraction_metadata: Extraction context (template prefix, taxonomy axis)
            config: Visualization configuration
            verbose: Enable verbose output
        """
        self.clustering_metadata = clustering_metadata
        self.code_assigned_results = code_assigned_results
        self.theme_enriched_codebook = theme_enriched_codebook
        self.extraction_metadata = extraction_metadata
        self.config = config or ExportVisualizerConfig()
        self.verbose = verbose
        self.verbose_reporter = VerboseReporter(verbose)

        # Initialize spaCy model for lemmatization
        self._nlp = None
        if SPACY_AVAILABLE:
            try:
                self._nlp = spacy.load(self.config.spacy_model)
            except OSError:
                if verbose:
                    print(f"Warning: spaCy model '{self.config.spacy_model}' not found. Trying 'en_core_web_sm'...")
                try:
                    self._nlp = spacy.load('en_core_web_sm')
                except OSError:
                    print("Warning: No spaCy model available. Lemmatization disabled.")

        # Build data structures for visualization
        self._theme_to_codes = self._build_theme_code_mapping()
        self._code_to_taxonomy_phrases = self._build_code_taxonomy_mapping()
        self._code_to_clusters = self._build_code_cluster_mapping()
        self._co_occurrence_matrix = self._build_co_occurrence_matrix()

        # Temp directory for intermediate files
        self._temp_dir = self.config.temp_dir or tempfile.mkdtemp(prefix='export_viz_')
        Path(self._temp_dir).mkdir(parents=True, exist_ok=True)

    def _build_theme_code_mapping(self) -> Dict[str, List[Dict[str, Any]]]:
        """Build mapping from theme -> list of codes with metadata."""
        theme_map = defaultdict(list)

        for entry in self.theme_enriched_codebook.codes:
            theme = entry.theme or 'Uncategorized'
            theme_map[theme].append({
                'code': entry.code,
                'definition': entry.definition,
                'category': entry.category or '',
                'source_cluster': entry.source_cluster
            })

        return dict(theme_map)

    def _build_code_taxonomy_mapping(self) -> Dict[str, List[str]]:
        """Build mapping from code -> list of taxonomy phrases from assigned ideas."""
        code_phrases = defaultdict(list)

        for result in self.code_assigned_results:
            if result.response_ideas:
                for idea in result.response_ideas:
                    if idea.assigned_codes and idea.taxonomy_phrase:
                        for code in idea.assigned_codes:
                            code_phrases[code].append(idea.taxonomy_phrase)

        return dict(code_phrases)

    def _build_code_cluster_mapping(self) -> Dict[str, List[str]]:
        """Build mapping from code -> source cluster IDs."""
        code_clusters = {}

        for entry in self.theme_enriched_codebook.codes:
            if entry.source_cluster:
                # Handle comma-separated cluster IDs
                if isinstance(entry.source_cluster, str):
                    clusters = [c.strip() for c in entry.source_cluster.split(',')]
                else:
                    clusters = [str(entry.source_cluster)]
                code_clusters[entry.code] = clusters

        return code_clusters

    def _build_co_occurrence_matrix(self) -> Dict[Tuple[str, str], int]:
        """Build co-occurrence counts for codes appearing on same response."""
        co_occurrence = defaultdict(int)

        for result in self.code_assigned_results:
            if result.response_ideas:
                # Collect all codes assigned to ideas in this response
                response_codes = set()
                for idea in result.response_ideas:
                    if idea.assigned_codes:
                        response_codes.update(idea.assigned_codes)

                # Count co-occurrences (pairs)
                codes_list = sorted(response_codes)
                for i, code1 in enumerate(codes_list):
                    for code2 in codes_list[i+1:]:
                        co_occurrence[(code1, code2)] += 1

        return dict(co_occurrence)

    def _lemmatize_phrases(self, phrases: List[str], pos_filter: str = 'all') -> List[str]:
        """
        Lemmatize phrases with optional POS filtering.

        Args:
            phrases: List of phrases to lemmatize
            pos_filter: 'all', 'nouns', 'adj_noun', 'bigrams'

        Returns:
            List of lemmatized words/phrases
        """
        if not self._nlp or not phrases:
            # Fallback: just return lowercase words
            words = []
            for phrase in phrases:
                words.extend(phrase.lower().split())
            return words

        result = []

        for phrase in phrases:
            doc = self._nlp(phrase.lower())

            if pos_filter == 'nouns':
                # Only nouns and proper nouns
                tokens = [token.lemma_ for token in doc if token.pos_ in ('NOUN', 'PROPN')]
            elif pos_filter == 'adj_noun':
                # Adjectives and nouns
                tokens = [token.lemma_ for token in doc if token.pos_ in ('ADJ', 'NOUN', 'PROPN')]
            elif pos_filter == 'bigrams':
                # Generate bigrams from all content words
                content_tokens = [token.lemma_ for token in doc if token.pos_ in ('ADJ', 'NOUN', 'PROPN', 'VERB')]
                tokens = []
                for i in range(len(content_tokens) - 1):
                    tokens.append(f"{content_tokens[i]} {content_tokens[i+1]}")
            else:  # 'all'
                # All lemmatized tokens except stopwords and punctuation
                tokens = [token.lemma_ for token in doc
                         if not token.is_stop and not token.is_punct and len(token.text) > 1]

            result.extend(tokens)

        return result

    def generate_dendrogram(self) -> Optional[Path]:
        """
        Generate hierarchical dendrogram visualization.

        Returns:
            Path to generated PNG file, or None if generation failed
        """
        self.verbose_reporter.step_start("Dendrogram", "Generating hierarchical tree visualization")

        if not self._theme_to_codes:
            print("Warning: No theme-code mappings available for dendrogram")
            return None

        fig, ax = plt.subplots(figsize=self.config.dendrogram_figsize)

        # Build hierarchy data
        themes = list(self._theme_to_codes.keys())
        n_themes = len(themes)

        # Assign colors to themes
        theme_colors = {}
        for i, theme in enumerate(themes):
            theme_colors[theme] = self.config.theme_colors[i % len(self.config.theme_colors)]

        # Calculate layout
        y_positions = {}  # Track y position for each node
        current_y = 0
        x_theme = 0.1
        x_code = 0.4
        x_cluster = 0.7

        # Draw hierarchy
        for theme_idx, theme in enumerate(themes):
            codes = self._theme_to_codes[theme]
            theme_color = theme_colors[theme]

            # Calculate theme y position (center of its codes)
            theme_y_start = current_y

            for code_idx, code_data in enumerate(codes[:self.config.dendrogram_max_items_per_level]):
                code = code_data['code']
                source_cluster = code_data.get('source_cluster', '')

                # Code position
                code_y = current_y
                y_positions[code] = code_y

                # Draw code node
                ax.scatter([x_code], [code_y], s=100, c=[theme_color], zorder=3)

                # Truncate long code names
                code_label = code[:30] + '...' if len(code) > 30 else code
                ax.text(x_code + 0.02, code_y, code_label, va='center', fontsize=8)

                # Draw clusters if available
                if source_cluster:
                    clusters = [c.strip() for c in str(source_cluster).split(',')]
                    for cl_idx, cluster in enumerate(clusters[:3]):  # Max 3 clusters per code
                        cluster_y = code_y + (cl_idx - len(clusters[:3])/2 + 0.5) * 0.3
                        ax.scatter([x_cluster], [cluster_y], s=50, c='gray', alpha=0.6, zorder=2)
                        ax.text(x_cluster + 0.02, cluster_y, f"C{cluster}", va='center', fontsize=6, color='gray')
                        # Connect code to cluster
                        ax.plot([x_code + 0.01, x_cluster - 0.01], [code_y, cluster_y],
                               'gray', alpha=0.3, linewidth=0.5)

                current_y += 1

            # Theme position (center of its codes)
            theme_y = (theme_y_start + current_y - 1) / 2
            y_positions[theme] = theme_y

            # Draw theme node
            ax.scatter([x_theme], [theme_y], s=300, c=[theme_color], zorder=4, marker='s')

            # Truncate theme names
            theme_label = theme[:25] + '...' if len(theme) > 25 else theme
            ax.text(x_theme - 0.08, theme_y, theme_label, va='center', ha='right',
                   fontsize=10, fontweight='bold', color=theme_color)

            # Connect theme to codes
            for code_data in codes[:self.config.dendrogram_max_items_per_level]:
                code = code_data['code']
                ax.plot([x_theme + 0.02, x_code - 0.02], [theme_y, y_positions[code]],
                       color=theme_color, alpha=0.5, linewidth=1)

            # Add spacing between themes
            current_y += 1

        # Styling
        ax.set_xlim(-0.1, 1.0)
        ax.set_ylim(-1, current_y)
        ax.axis('off')

        # Add column headers
        ax.text(x_theme, current_y + 0.5, 'THEMES', ha='center', fontsize=12, fontweight='bold')
        ax.text(x_code, current_y + 0.5, 'CODES', ha='center', fontsize=12, fontweight='bold')
        ax.text(x_cluster, current_y + 0.5, 'CLUSTERS', ha='center', fontsize=12, fontweight='bold')

        # Title
        title = "Codebook Hierarchy: Theme → Code → Cluster"
        if self.extraction_metadata and self.extraction_metadata.var_lab:
            title += f"\n{self.extraction_metadata.var_lab[:80]}"
        ax.set_title(title, fontsize=14, fontweight='bold', pad=20)

        # Save
        output_path = Path(self._temp_dir) / 'dendrogram.png'
        plt.tight_layout()
        plt.savefig(output_path, dpi=self.config.dendrogram_dpi, bbox_inches='tight',
                   facecolor='white', edgecolor='none')
        plt.close(fig)

        self.verbose_reporter.step_complete(f"Dendrogram saved: {output_path}")
        return output_path

    def generate_word_cloud_grid(self) -> Optional[Path]:
        """
        Generate grid of word clouds for all codes.

        Returns:
            Path to generated PNG file, or None if generation failed
        """
        if not WORDCLOUD_AVAILABLE:
            print("Warning: wordcloud library not available. Skipping word cloud generation.")
            return None

        self.verbose_reporter.step_start("Word Clouds", "Generating word cloud grid")

        codes = list(self._code_to_taxonomy_phrases.keys())
        if not codes:
            print("Warning: No code-taxonomy mappings available for word clouds")
            return None

        n_codes = len(codes)
        n_cols = min(self.config.wordcloud_grid_cols, n_codes)
        n_rows = (n_codes + n_cols - 1) // n_cols

        # Calculate figure size
        fig_width = n_cols * (self.config.wordcloud_cell_width / 100)
        fig_height = n_rows * (self.config.wordcloud_cell_height / 100 + 0.5)  # Extra space for labels

        fig, axes = plt.subplots(n_rows, n_cols, figsize=(fig_width, fig_height))

        # Flatten axes for easy iteration
        if n_rows == 1 and n_cols == 1:
            axes = [[axes]]
        elif n_rows == 1:
            axes = [axes]
        elif n_cols == 1:
            axes = [[ax] for ax in axes]

        # Get theme colors for codes
        code_to_theme = self.theme_enriched_codebook.code_to_theme_mapping or {}
        theme_colors = {}
        themes = list(set(code_to_theme.values()))
        for i, theme in enumerate(themes):
            theme_colors[theme] = self.config.theme_colors[i % len(self.config.theme_colors)]

        for idx, code in enumerate(codes):
            row = idx // n_cols
            col = idx % n_cols
            ax = axes[row][col]

            phrases = self._code_to_taxonomy_phrases.get(code, [])

            if phrases:
                # Lemmatize and process
                words = self._lemmatize_phrases(phrases, self.config.pos_filter)

                if words:
                    # Create word frequency dict
                    word_freq = defaultdict(int)
                    for word in words:
                        word_freq[word] += 1

                    # Generate word cloud
                    try:
                        # Get color based on theme
                        theme = code_to_theme.get(code, 'default')
                        color = theme_colors.get(theme, '#1f77b4')

                        wc = WordCloud(
                            width=self.config.wordcloud_cell_width,
                            height=self.config.wordcloud_cell_height,
                            max_words=self.config.wordcloud_max_words,
                            background_color=self.config.wordcloud_background,
                            colormap=self.config.wordcloud_colormap,
                            prefer_horizontal=0.7,
                            min_font_size=8
                        ).generate_from_frequencies(word_freq)

                        ax.imshow(wc, interpolation='bilinear')
                    except Exception as e:
                        ax.text(0.5, 0.5, f"Error:\n{str(e)[:20]}",
                               ha='center', va='center', fontsize=8)
                else:
                    ax.text(0.5, 0.5, "No words", ha='center', va='center', fontsize=10, color='gray')
            else:
                ax.text(0.5, 0.5, "No data", ha='center', va='center', fontsize=10, color='gray')

            ax.axis('off')

            # Add code label above
            code_label = code[:25] + '...' if len(code) > 25 else code
            ax.set_title(code_label, fontsize=9, fontweight='bold', pad=5)

        # Hide empty subplots
        for idx in range(n_codes, n_rows * n_cols):
            row = idx // n_cols
            col = idx % n_cols
            axes[row][col].axis('off')

        # Main title
        plt.suptitle("Word Clouds by Code (based on taxonomy phrases)",
                    fontsize=14, fontweight='bold', y=1.02)

        # Save
        output_path = Path(self._temp_dir) / 'wordcloud_grid.png'
        plt.tight_layout()
        plt.savefig(output_path, dpi=self.config.wordcloud_dpi, bbox_inches='tight',
                   facecolor='white', edgecolor='none')
        plt.close(fig)

        self.verbose_reporter.step_complete(f"Word cloud grid saved: {output_path}")
        return output_path

    def generate_network_html(self, output_dir: Optional[str] = None) -> Optional[Path]:
        """
        Generate interactive network graph as HTML file.

        Args:
            output_dir: Directory for output file (defaults to temp_dir)

        Returns:
            Path to generated HTML file, or None if generation failed
        """
        if not NETWORKX_AVAILABLE or not PLOTLY_AVAILABLE:
            print("Warning: networkx or plotly not available. Skipping network graph.")
            return None

        self.verbose_reporter.step_start("Network Graph", "Generating interactive network visualization")

        # Build network graph
        G = nx.Graph()

        # Get theme colors
        themes = list(self._theme_to_codes.keys())
        theme_colors = {}
        for i, theme in enumerate(themes):
            theme_colors[theme] = self.config.theme_colors[i % len(self.config.theme_colors)]

        # Add theme nodes
        for theme in themes:
            G.add_node(f"theme:{theme}",
                      node_type='theme',
                      label=theme,
                      color=theme_colors[theme],
                      size=self.config.network_theme_size)

        # Add code nodes and theme-code edges
        code_to_theme = self.theme_enriched_codebook.code_to_theme_mapping or {}
        for code, theme in code_to_theme.items():
            color = theme_colors.get(theme, '#666666')
            G.add_node(f"code:{code}",
                      node_type='code',
                      label=code,
                      color=color,
                      size=self.config.network_code_size)
            G.add_edge(f"theme:{theme}", f"code:{code}",
                      edge_type='hierarchy', weight=2)

        # Add cluster nodes and code-cluster edges
        all_clusters = set()
        for code, clusters in self._code_to_clusters.items():
            for cluster in clusters:
                cluster_id = f"cluster:{cluster}"
                if cluster_id not in G:
                    G.add_node(cluster_id,
                              node_type='cluster',
                              label=f"C{cluster}",
                              color='#cccccc',
                              size=self.config.network_cluster_size)
                    all_clusters.add(cluster_id)
                G.add_edge(f"code:{code}", cluster_id,
                          edge_type='source', weight=1)

        # Add code co-occurrence edges
        for (code1, code2), count in self._co_occurrence_matrix.items():
            if count >= 2:  # Only show if co-occur at least twice
                G.add_edge(f"code:{code1}", f"code:{code2}",
                          edge_type='cooccurrence', weight=count)

        # Apply layout
        if self.config.network_layout == 'kamada_kawai':
            pos = nx.kamada_kawai_layout(G)
        elif self.config.network_layout == 'circular':
            pos = nx.circular_layout(G)
        else:  # spring
            pos = nx.spring_layout(G, k=2, iterations=50)

        # Create Plotly traces
        # Edge traces (by type)
        edge_traces = []

        for edge_type, style in [('hierarchy', dict(color='rgba(100,100,100,0.5)', width=2, dash='solid')),
                                  ('source', dict(color='rgba(150,150,150,0.3)', width=1, dash='dot')),
                                  ('cooccurrence', dict(color='rgba(255,165,0,0.4)', width=1.5, dash='dash'))]:
            edge_x = []
            edge_y = []
            for edge in G.edges(data=True):
                if edge[2].get('edge_type') == edge_type:
                    x0, y0 = pos[edge[0]]
                    x1, y1 = pos[edge[1]]
                    edge_x.extend([x0, x1, None])
                    edge_y.extend([y0, y1, None])

            if edge_x:
                edge_traces.append(go.Scatter(
                    x=edge_x, y=edge_y,
                    mode='lines',
                    line=style,
                    hoverinfo='none',
                    name=edge_type.capitalize()
                ))

        # Node traces (by type)
        node_traces = []

        for node_type in ['theme', 'code', 'cluster']:
            node_x = []
            node_y = []
            node_text = []
            node_color = []
            node_size = []

            for node, data in G.nodes(data=True):
                if data.get('node_type') == node_type:
                    x, y = pos[node]
                    node_x.append(x)
                    node_y.append(y)
                    node_text.append(data.get('label', node))
                    node_color.append(data.get('color', '#666666'))
                    node_size.append(data.get('size', 20))

            if node_x:
                node_traces.append(go.Scatter(
                    x=node_x, y=node_y,
                    mode='markers+text' if node_type != 'cluster' else 'markers',
                    text=node_text,
                    textposition='top center' if node_type == 'theme' else 'bottom center',
                    textfont=dict(size=12 if node_type == 'theme' else 8),
                    marker=dict(
                        size=node_size,
                        color=node_color,
                        line=dict(width=2, color='white'),
                        symbol='square' if node_type == 'theme' else 'circle'
                    ),
                    hovertext=node_text,
                    hoverinfo='text',
                    name=node_type.capitalize() + 's'
                ))

        # Create figure
        fig = go.Figure(
            data=edge_traces + node_traces,
            layout=go.Layout(
                title=dict(
                    text="Theme-Code-Cluster Network",
                    x=0.5,
                    font=dict(size=16)
                ),
                showlegend=True,
                legend=dict(x=1.02, y=0.5),
                hovermode='closest',
                width=self.config.network_figsize[0],
                height=self.config.network_figsize[1],
                xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                plot_bgcolor='white',
                paper_bgcolor='white'
            )
        )

        # Add annotation for legend
        fig.add_annotation(
            x=1.15, y=0.95,
            xref='paper', yref='paper',
            text="<b>Legend:</b><br>Squares = Themes<br>Large circles = Codes<br>Small circles = Clusters<br>Solid = Hierarchy<br>Dotted = Source<br>Dashed = Co-occurrence",
            showarrow=False,
            font=dict(size=10),
            align='left',
            bgcolor='rgba(255,255,255,0.8)',
            bordercolor='gray',
            borderwidth=1
        )

        # Save
        output_dir = output_dir or self._temp_dir
        output_path = Path(output_dir) / 'network_graph.html'
        fig.write_html(str(output_path), include_plotlyjs='cdn')

        self.verbose_reporter.step_complete(f"Network graph saved: {output_path}")
        return output_path

    def add_visualizations_to_workbook(self, workbook: Workbook) -> Workbook:
        """
        Add visualization sheets to existing Excel workbook.

        Args:
            workbook: openpyxl Workbook to add sheets to

        Returns:
            Modified workbook with visualization sheets
        """
        self.verbose_reporter.section_header("ADDING VISUALIZATIONS TO EXCEL")

        # Generate dendrogram
        dendrogram_path = self.generate_dendrogram()
        if dendrogram_path and dendrogram_path.exists():
            ws = workbook.create_sheet(title="Dendrogram")
            img = ExcelImage(str(dendrogram_path))
            # Scale to fit nicely
            img.width = 1200
            img.height = 800
            ws.add_image(img, 'A1')
            self.verbose_reporter.stat_line("Added Dendrogram sheet")

        # Generate word cloud grid
        wordcloud_path = self.generate_word_cloud_grid()
        if wordcloud_path and wordcloud_path.exists():
            ws = workbook.create_sheet(title="Word Clouds")
            img = ExcelImage(str(wordcloud_path))
            # Scale based on grid size
            img.width = min(1400, len(self._code_to_taxonomy_phrases) * 150)
            img.height = min(1000, (len(self._code_to_taxonomy_phrases) // 4 + 1) * 200)
            ws.add_image(img, 'A1')
            self.verbose_reporter.stat_line("Added Word Clouds sheet")

        return workbook

    def generate_all(self, output_dir: str) -> Dict[str, Optional[Path]]:
        """
        Generate all visualizations.

        Args:
            output_dir: Directory for output files

        Returns:
            Dict mapping visualization name to output path
        """
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        results = {
            'dendrogram': self.generate_dendrogram(),
            'wordcloud_grid': self.generate_word_cloud_grid(),
            'network_html': self.generate_network_html(output_dir)
        }

        # Copy temp files to output dir if different
        if self._temp_dir != output_dir:
            import shutil
            for name, path in results.items():
                if path and path.exists() and str(path.parent) == self._temp_dir:
                    new_path = Path(output_dir) / path.name
                    shutil.copy(path, new_path)
                    results[name] = new_path

        return results
