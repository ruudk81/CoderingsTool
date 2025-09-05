import os, sys; sys.path.extend([p for p in [os.getcwd().split('coderingsTool')[0] + suffix for suffix in ['', 'coderingsTool', 'coderingsTool/src', 'coderingsTool/src/utils']] if p not in sys.path]) if 'coderingsTool' in os.getcwd() else None

# === MODULES ========================================================================================================
import os
from typing import List, Dict, Optional
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# === MODELS ========================================================================================================
import models

# === UTILS ========================================================================================================
from .verboseReporter import VerboseReporter
from .codeGenerator import CodeGeneratorReasoningResults

class CodeAssignmentExporter:
    
    def __init__(self, verbose: bool = True):
        self.verbose = verbose
        self.verbose_reporter = VerboseReporter(verbose, capture_logging=True)
    
    def _sanitize_excel_text(self, text):
        """Remove or replace control characters that Excel cannot handle"""
        if text is None or not isinstance(text, str):
            return text
        
        # Remove all control characters (0x00-0x1F) except tab, newline, carriage return
        # These are characters that openpyxl/Excel cannot handle
        sanitized = ''.join(char if ord(char) > 31 or char in '\t\n\r' else ' ' for char in text)
        
        # Also handle the specific problematic character we saw (0x08 - backspace)
        sanitized = sanitized.replace('\x08', '')
        
        return sanitized
        
    def export_to_excel(self,
                       code_assigned_results: List[models.CodeAssignedModel],
                       theme_enriched_codebook: models.ThemeEnrichedCodebookModel,
                       filename: str,
                       var_name: str,
                       export_dir: Optional[str] = None) -> str:
        """
        Export code assignment results to Excel with all requested columns:
        - respondent_id
        - idea_id
        - initial_cluster_id (parent cluster)
        - source_cluster_id (sub-cluster)
        - code_label
        - code_description
        - assignment_rationale
        - theme_name
        - theme_description
        """
        
        self.verbose_reporter.section_header("EXPORTING CODE ASSIGNMENTS TO EXCEL")
        
        # Prepare data for export
        export_data = []
        
        # Create comprehensive code mapping from enriched codebook
        code_to_info = {}
        for code_entry in theme_enriched_codebook.codes:
            code_to_info[code_entry.code] = {
                'theme_name': code_entry.theme or 'No Theme',
                'theme_description': code_entry.theme_description or 'No Description',
                'code_description': code_entry.definition,
                'source_cluster': code_entry.source_cluster  # This is the sub-cluster ID like "12-1"
            }
        
        # Process each result
        for result in code_assigned_results:
            respondent_id = result.respondent_id
            
            if result.response_ideas:
                for idea in result.response_ideas:
                    # Process each assigned code
                    if idea.assigned_codes:
                        for code in idea.assigned_codes:
                            code_info = code_to_info.get(code, {
                                'theme_name': 'Unknown Theme',
                                'theme_description': 'Unknown Description',
                                'code_description': 'Unknown Code',
                                'source_cluster': None
                            })
                            
                            # Get source cluster info from the enriched codebook
                            source_cluster = code_info['source_cluster']
                            
                            # Extract parent cluster from source cluster
                            if source_cluster and isinstance(source_cluster, str) and '-' in source_cluster:
                                parent_cluster = source_cluster.split('-')[0]
                            else:
                                parent_cluster = source_cluster
                            
                            row_data = {
                                'respondent_id': respondent_id,
                                'original_response': result.response,
                                'idea_id': idea.idea_id,
                                'idea_text': idea.idea,
                                'initial_cluster_id': parent_cluster,
                                'source_cluster_id': source_cluster,  # This is the sub-cluster from enriched codebook
                                'code_label': code,
                                'code_description': code_info['code_description'],
                                'assignment_rationale': idea.assignment_rationale if hasattr(idea, 'assignment_rationale') else '',
                                'assignment_confidence': idea.assignment_confidence if hasattr(idea, 'assignment_confidence') else None,
                                'theme_name': code_info['theme_name'],
                                'theme_description': code_info['theme_description']
                            }
                            export_data.append(row_data)
                    else:
                        # No codes assigned - still export the row with empty code fields
                        # Get cluster info from the idea itself as fallback
                        initial_cluster = idea.initial_cluster if hasattr(idea, 'initial_cluster') else None
                        if initial_cluster and isinstance(initial_cluster, str) and '-' in initial_cluster:
                            parent_cluster = initial_cluster.split('-')[0]
                            source_cluster = initial_cluster
                        else:
                            parent_cluster = initial_cluster
                            source_cluster = initial_cluster
                            
                        row_data = {
                            'respondent_id': respondent_id,
                            'original_response': result.response,
                            'idea_id': idea.idea_id,
                            'idea_text': idea.idea,
                            'initial_cluster_id': parent_cluster,
                            'source_cluster_id': source_cluster,
                            'code_label': 'No Code Assigned',
                            'code_description': '',
                            'assignment_rationale': idea.assignment_rationale if hasattr(idea, 'assignment_rationale') else '',
                            'assignment_confidence': idea.assignment_confidence if hasattr(idea, 'assignment_confidence') else None,
                            'theme_name': '',
                            'theme_description': ''
                        }
                        export_data.append(row_data)
        
        # Convert to DataFrame
        df = pd.DataFrame(export_data)
        
        # Create export directory if not provided
        if export_dir is None:
            # Get project root (3 levels up from utils folder)
            project_root = Path(__file__).parent.parent.parent
            export_dir = os.path.join(project_root, 'exports')
        
        Path(export_dir).mkdir(parents=True, exist_ok=True)
        
        # Create output filename
        base_name = Path(filename).stem
        output_filename = f"{base_name}_{var_name}_code_assignments.xlsx"
        output_path = os.path.join(export_dir, output_filename)
        
        print(f"About to export DataFrame with shape: {df.shape}")
        #print(f"DataFrame columns: {list(df.columns)}")
        # if len(df) > 0:
        #     print(f"Sample data - original_response: {str(df.iloc[0]['original_response'])[:100]}...")
        #     print(f"Sample data - idea_text: {str(df.iloc[0]['idea_text'])[:100]}...")
        
        # Export to Excel with formatting
        self._write_formatted_excel(df, output_path, var_name)
        
        # Report statistics
        self.verbose_reporter.stat_line(f"Total rows exported: {len(export_data)}")
        self.verbose_reporter.stat_line(f"Unique respondents: {df['respondent_id'].nunique()}")
        self.verbose_reporter.stat_line(f"Unique ideas: {df['idea_id'].nunique()}")
        self.verbose_reporter.stat_line(f"Unique codes assigned: {df[df['code_label'] != 'No Code Assigned']['code_label'].nunique()}")
        
        return output_path
    
    def export_to_excel_with_reasoning(self,
                                      code_assigned_results: List[models.CodeAssignedModel],
                                      theme_enriched_codebook: models.ThemeEnrichedCodebookModel,
                                      reasoning_results: CodeGeneratorReasoningResults,
                                      filename: str,
                                      var_name: str,
                                      export_dir: Optional[str] = None) -> str:
        """
        Export code assignment results with step 7 reasoning data to Excel.
        Includes all data from regular export plus step 3 and step 4 reasoning.
        """
        
        self.verbose_reporter.section_header("EXPORTING CODE ASSIGNMENTS WITH REASONING TO EXCEL")
        
        # Prepare data for export
        export_data = []
        
        # Create comprehensive code mapping from enriched codebook
        code_to_info = {}
        for code_entry in theme_enriched_codebook.codes:
            code_to_info[code_entry.code] = {
                'theme_name': code_entry.theme or 'No Theme',
                'theme_description': code_entry.theme_description or 'No Description',
                'code_description': code_entry.definition,
                'source_cluster': code_entry.source_cluster  # This is the sub-cluster ID like "12-1"
            }
        
        # Create reasoning mapping from step 7 data
        reasoning_mapping = self._create_reasoning_mapping(reasoning_results)
        
        # Process each result
        for result in code_assigned_results:
            respondent_id = result.respondent_id
            
            if result.response_ideas:
                for idea in result.response_ideas:
                    # Process each assigned code
                    if idea.assigned_codes:
                        for code in idea.assigned_codes:
                            code_info = code_to_info.get(code, {
                                'theme_name': 'Unknown Theme',
                                'theme_description': 'Unknown Description',
                                'code_description': 'Unknown Code',
                                'source_cluster': None
                            })
                            
                            # Get source cluster info from the enriched codebook
                            source_cluster = code_info['source_cluster']
                            
                            # Extract parent cluster from source cluster
                            if source_cluster and isinstance(source_cluster, str) and '-' in source_cluster:
                                parent_cluster = source_cluster.split('-')[0]
                            else:
                                parent_cluster = source_cluster
                            
                            # Get reasoning data for this cluster
                            reasoning_data = self._get_reasoning_for_cluster(reasoning_mapping, source_cluster, parent_cluster)
                            
                            row_data = {
                                'respondent_id': respondent_id,
                                'original_response': result.response,
                                'idea_id': idea.idea_id,
                                'idea_text': idea.idea,
                                'initial_cluster_id': parent_cluster,
                                'source_cluster_id': source_cluster,
                                'code_label': code,
                                'code_description': code_info['code_description'],
                                'assignment_rationale': idea.assignment_rationale if hasattr(idea, 'assignment_rationale') else '',
                                'assignment_confidence': idea.assignment_confidence if hasattr(idea, 'assignment_confidence') else None,
                                'theme_name': code_info['theme_name'],
                                'theme_description': code_info['theme_description'],
                                # Step 7 reasoning data
                                'codegen_theme': reasoning_data.get('codegen_theme', ''),
                                'codegen_recommendation': reasoning_data.get('codegen_recommendation', ''),
                                'codebook_validation': reasoning_data.get('codebook_validation', '')
                            }
                            export_data.append(row_data)
                    else:
                        # No codes assigned - still export the row with empty code fields
                        initial_cluster = idea.initial_cluster if hasattr(idea, 'initial_cluster') else None
                        if initial_cluster and isinstance(initial_cluster, str) and '-' in initial_cluster:
                            parent_cluster = initial_cluster.split('-')[0]
                            source_cluster = initial_cluster
                        else:
                            parent_cluster = initial_cluster
                            source_cluster = initial_cluster
                        
                        # Get reasoning data for this cluster
                        reasoning_data = self._get_reasoning_for_cluster(reasoning_mapping, source_cluster, parent_cluster)
                            
                        row_data = {
                            'respondent_id': respondent_id,
                            'original_response': result.response,
                            'idea_id': idea.idea_id,
                            'idea_text': idea.idea,
                            'initial_cluster_id': parent_cluster,
                            'source_cluster_id': source_cluster,
                            'code_label': 'No Code Assigned',
                            'code_description': '',
                            'assignment_rationale': idea.assignment_rationale if hasattr(idea, 'assignment_rationale') else '',
                            'assignment_confidence': idea.assignment_confidence if hasattr(idea, 'assignment_confidence') else None,
                            'theme_name': '',
                            'theme_description': '',
                            # Step 7 reasoning data
                            'codegen_theme': reasoning_data.get('codegen_theme', ''),
                            'codegen_recommendation': reasoning_data.get('codegen_recommendation', ''),
                            'codebook_validation': reasoning_data.get('codebook_validation', '')
                        }
                        export_data.append(row_data)
        
        # Convert to DataFrame
        df = pd.DataFrame(export_data)
        
        # Create export directory if not provided
        if export_dir is None:
            # Get project root (3 levels up from utils folder)
            project_root = Path(__file__).parent.parent.parent
            export_dir = os.path.join(project_root, 'exports')
        
        Path(export_dir).mkdir(parents=True, exist_ok=True)
        
        # Create output filename
        base_name = Path(filename).stem
        output_filename = f"{base_name}_{var_name}_code_assignments_with_reasoning.xlsx"
        output_path = os.path.join(export_dir, output_filename)
        
        # Export to Excel with formatting
        self._write_formatted_excel_with_reasoning(df, output_path, var_name)
        
        # Report statistics
        self.verbose_reporter.stat_line(f"Total rows exported: {len(export_data)}")
        self.verbose_reporter.stat_line(f"Unique respondents: {df['respondent_id'].nunique()}")
        self.verbose_reporter.stat_line(f"Unique ideas: {df['idea_id'].nunique()}")
        self.verbose_reporter.stat_line(f"Unique codes assigned: {df[df['code_label'] != 'No Code Assigned']['code_label'].nunique()}")
        self.verbose_reporter.stat_line(f"Excel file with reasoning saved: {output_path}")
        
        return output_path
    
    def _write_formatted_excel(self, df: pd.DataFrame, output_path: str, var_name: str):
        """Write DataFrame to Excel with formatting"""
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "codering"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        headers = [
            'Respondent ID',
            'Original Response',
            'Idea ID',
            'Idea Text',
            'Initial Cluster ID',
            'Source Cluster ID',
            'Code Label',
            'Code Description',
            'Assignment Rationale',
            'Assignment Confidence',
            'Theme Name',
            'Theme Description'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        try:
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
                 for c_idx, value in enumerate(row, 1):
                    # Handle None values for Excel
                    if value is None:
                        value = ""
                    
                    # Sanitize string values to remove control characters
                    if isinstance(value, str):
                        value = self._sanitize_excel_text(value)    
                    try:
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = border
                    except Exception as cell_error:
                        print(f"DEBUG: Error writing cell at row {r_idx}, col {c_idx}")
                        print(f"DEBUG: Cell value: {str(value)[:200]}...")
                        print(f"DEBUG: Cell error: {cell_error}")
                        raise
                    
                    # Format confidence values
                    if c_idx == 10 and value is not None:  # Assignment confidence column (shifted by 1 due to new column)
                        try:
                            cell.value = float(value)
                            cell.number_format = '0.00'
                        except:
                            pass
        except Exception as e:
            print(f"DEBUG: Error during data writing: {e}")
            raise
        
        # Adjust column widths
        column_widths = {
            'A': 15,  # Respondent ID
            'B': 50,  # Original Response
            'C': 15,  # Idea ID
            'D': 50,  # Idea Text
            'E': 18,  # Initial Cluster ID
            'F': 18,  # Source Cluster ID
            'G': 30,  # Code Label
            'H': 50,  # Code Description
            'I': 60,  # Assignment Rationale
            'J': 20,  # Assignment Confidence
            'K': 30,  # Theme Name
            'L': 50   # Theme Description
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze the header row
        ws.freeze_panes = 'A2'
        
        # Add summary sheet
        summary_ws = wb.create_sheet(title="Summary")
        
        # Summary statistics
        summary_data = [
            ["Summary Statistics", ""],
            ["", ""],
            ["Total Assignments", len(df)],
            ["Unique Respondents", df['respondent_id'].nunique()],
            ["Unique Ideas", df['idea_id'].nunique()],
            ["Unique Codes", df[df['code_label'] != 'No Code Assigned']['code_label'].nunique()],
            ["Unique Themes", df[df['theme_name'] != '']['theme_name'].nunique()],
            ["", ""],
            ["Code Frequency", "Count"],
        ]
        
        # Add code frequency
        code_freq = df[df['code_label'] != 'No Code Assigned']['code_label'].value_counts()
        for code, count in code_freq.items():
            summary_data.append([code, count])
        
        # Write summary data
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1 or (row_idx == 9 and col_idx == 1):
                    cell.font = Font(bold=True, size=14)
                elif row_idx in [3, 4, 5, 6, 7]:
                    if col_idx == 1:
                        cell.font = Font(bold=True)
        
        # Adjust summary column widths
        summary_ws.column_dimensions['A'].width = 30
        summary_ws.column_dimensions['B'].width = 15
        
        # Save workbook
        try:
            wb.save(output_path)
        except Exception as e:
            print(f"DEBUG: Error saving workbook: {e}")
            raise
    
    def _create_reasoning_mapping(self, reasoning_results: CodeGeneratorReasoningResults) -> Dict[str, Dict[str, str]]:
        """Create a mapping from cluster IDs to reasoning data"""
        reasoning_mapping = {}
        
        # Process step 1 summaries (theme analysis)
        for cluster_id, step1_data in reasoning_results.step1_summaries.items():
            cluster_key = str(cluster_id)
            if cluster_key not in reasoning_mapping:
                reasoning_mapping[cluster_key] = {}
            
            # Extract analysis from step1 data for Codegen_theme
            analysis = step1_data.get('analysis', '')
            reasoning_mapping[cluster_key]['codegen_theme'] = analysis
        
        # Process step 3 recommendations (code generation decisions)
        for cluster_id, step3_data in reasoning_results.step3_recommendations.items():
            cluster_key = str(cluster_id)
            if cluster_key not in reasoning_mapping:
                reasoning_mapping[cluster_key] = {}
            
            if 'coding_decisions' in step3_data and step3_data['coding_decisions']:
                # Get the first coding decision (there might be multiple themes per cluster)
                first_decision = step3_data['coding_decisions'][0]
                decision = first_decision.get('decision', '')
                justification = first_decision.get('justification', '')
                # Combine decision + justification for Codegen_recommendation
                combined_recommendation = f"{decision}: {justification}" if decision and justification else decision or justification
                reasoning_mapping[cluster_key]['codegen_recommendation'] = combined_recommendation
        
        # Process step 4 validations
        for cluster_id, step4_data in reasoning_results.step4_validations.items():
            cluster_key = str(cluster_id)
            if cluster_key not in reasoning_mapping:
                reasoning_mapping[cluster_key] = {}
            
            if 'code_validations' in step4_data and step4_data['code_validations']:
                # Get the first validation (there might be multiple validations per cluster)
                first_validation = step4_data['code_validations'][0]
                code_label = first_validation.get('code_label', '')
                justification = first_validation.get('decision_rationale', '')
                # Combine code label + justification for Codebook_validation
                combined_validation = f"{code_label}: {justification}" if code_label and justification else code_label or justification
                reasoning_mapping[cluster_key]['codebook_validation'] = combined_validation
        
        return reasoning_mapping
    
    def _get_reasoning_for_cluster(self, reasoning_mapping: Dict[str, Dict[str, str]], 
                                 source_cluster: Optional[str], 
                                 parent_cluster: Optional[str]) -> Dict[str, str]:
        """Get reasoning data for a specific cluster, trying source cluster first then parent cluster"""
        default_reasoning = {'codegen_theme': '', 'codegen_recommendation': '', 'codebook_validation': ''}
        
        # Try source cluster first (sub-cluster like "12-1")
        if source_cluster and str(source_cluster) in reasoning_mapping:
            return {**default_reasoning, **reasoning_mapping[str(source_cluster)]}
        
        # Try parent cluster (main cluster like "12")
        if parent_cluster and str(parent_cluster) in reasoning_mapping:
            return {**default_reasoning, **reasoning_mapping[str(parent_cluster)]}
        
        return default_reasoning
    
    def _write_formatted_excel_with_reasoning(self, df: pd.DataFrame, output_path: str, var_name: str):
        """Write DataFrame to Excel with formatting including reasoning columns"""
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "codering"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        headers = [
            'Respondent ID',
            'Original Response',
            'Idea ID',
            'Idea Text',
            'Initial Cluster ID',
            'Source Cluster ID',
            'Code Label',
            'Code Description',
            'Assignment Rationale',
            'Assignment Confidence',
            'Theme Name',
            'Theme Description',
            'Codegen_theme',
            'Codegen_recommendation',
            'Codebook_validation'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                # Handle None values for Excel
                if value is None:
                    value = ""
                
                # Sanitize string values to remove control characters
                if isinstance(value, str):
                    value = self._sanitize_excel_text(value)
                    # Also apply length limit to avoid openpyxl bugs
                    if len(value) > 255:  # Much shorter limit to avoid openpyxl bugs
                        value = value[:255] + "..."
                    
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                
                # Format confidence values
                if c_idx == 10 and value is not None:  # Assignment confidence column
                    try:
                        cell.value = float(value)
                        cell.number_format = '0.00'
                    except:
                        pass
        
        # Adjust column widths
        column_widths = {
            'A': 15,  # Respondent ID
            'B': 50,  # Original Response
            'C': 15,  # Idea ID
            'D': 50,  # Idea Text
            'E': 18,  # Initial Cluster ID
            'F': 18,  # Source Cluster ID
            'G': 30,  # Code Label
            'H': 50,  # Code Description
            'I': 60,  # Assignment Rationale
            'J': 20,  # Assignment Confidence
            'K': 30,  # Theme Name
            'L': 50,  # Theme Description
            'M': 60,  # Codegen_theme
            'N': 60,  # Codegen_recommendation
            'O': 60   # Codebook_validation
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze the header row
        ws.freeze_panes = 'A2'
        
        # Add summary sheet
        summary_ws = wb.create_sheet(title="Summary")
        
        # Summary statistics
        summary_data = [
            ["Summary Statistics", ""],
            ["", ""],
            ["Total Assignments", len(df)],
            ["Unique Respondents", df['respondent_id'].nunique()],
            ["Unique Ideas", df['idea_id'].nunique()],
            ["Unique Codes", df[df['code_label'] != 'No Code Assigned']['code_label'].nunique()],
            ["Unique Themes", df[df['theme_name'] != '']['theme_name'].nunique()],
            ["", ""],
            ["Codegen Recommendations", "Count"],
        ]
        
        # Add codegen recommendation frequency
        codegen_rec_freq = df[df['codegen_recommendation'] != '']['codegen_recommendation'].value_counts()
        for recommendation, count in codegen_rec_freq.items():
            summary_data.append([recommendation, count])
        
        summary_data.extend([
            ["", ""],
            ["Code Frequency", "Count"],
        ])
        
        # Add code frequency
        code_freq = df[df['code_label'] != 'No Code Assigned']['code_label'].value_counts()
        for code, count in code_freq.items():
            summary_data.append([code, count])
        
        # Write summary data
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1 or (row_idx == 9 and col_idx == 1) or ("Frequency" in str(value) and col_idx == 1):
                    cell.font = Font(bold=True, size=14)
                elif row_idx in [3, 4, 5, 6, 7]:
                    if col_idx == 1:
                        cell.font = Font(bold=True)
        
        # Adjust summary column widths
        summary_ws.column_dimensions['A'].width = 30
        summary_ws.column_dimensions['B'].width = 15
        
        # Save workbook
        try:
            wb.save(output_path)
        except Exception as e:
            print(f"DEBUG: Error saving workbook: {e}")
            raise