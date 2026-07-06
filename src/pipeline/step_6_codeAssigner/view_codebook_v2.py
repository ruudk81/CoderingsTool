"""Experimental v2 of the codebook — readability workbench (Taxonomie grof).

Isolated from the production exporter (`view_codebook.py`): it imports only the
DATA builders from it and never modifies it. Writes a separate `..._v2.xlsx`, so
the production workbook stays byte-for-byte identical. If this experiment
succeeds, `write_sheet_v2` gets folded back into `view_codebook.py` for all sheets
(app + CLI) as a single promotion PR.

Run (from src/):
    python -m pipeline.step_6_codeAssigner.view_codebook_v2

Data comes 1:1 from `view_codebook.build_domain_facet_attr` (single source of
truth); only the presentation layer is forked here.
"""
import sys
from pathlib import Path

src_dir = Path(__file__).parent.parent.parent
if str(src_dir) not in sys.path:
    sys.path.insert(0, str(src_dir))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from test_data import TEST_DATA
from pipeline.step_6_codeAssigner import view_codebook as vc


# =============================================================================
# STYLING KNOBS — the v2 experiment surface (iterated per instruction)
# =============================================================================
_HDR_FILL = PatternFill("solid", fgColor="366092")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_VAL_COLOR = {"+": "2E7D32", "-": "C62828", "~": "777777"}

# Readability tweak #1 — pull the hierarchy columns tight.
# Only one hierarchy column is filled per row, so a non-terminal level (domein,
# facet, …) can be narrow: its label overflows into the blank cells to its right
# and stays fully readable, while the tree reads as one compact block. The
# terminal level (attribuut) borders the filled metric columns, so it keeps an
# auto width. Widths are per non-terminal level; the list extends by repeating
# its last value. This is the knob we tune.
HIER_WIDTHS = [12, 18]   # domein, facet

# Readability tweak #2 — per-level bullet prefix on the hierarchy labels, so
# facets and attributes are visually marked and distinct from each other. Keyed
# by depth (0 = domein stays unmarked). Any Unicode glyph works; tune here.
BULLETS = {1: "• ", 2: "– "}   # facet = gevuld rondje, attribuut = streepje


def _within_parent_shares(rows):
    """Bruto share of each row within its parent: n(row) / n(parent). The parent
    is the nearest preceding row one level up (rows are a pre-order traversal:
    domain → its facets → each facet's attributes). Returns two aligned lists:
    `share` (fraction, or None for depth-0 / no parent) and `sole` (True when the
    parent has exactly one child — used to suppress a trivial inline '(100%)')."""
    n = len(rows)
    share = [None] * n
    parent_of = [None] * n
    last_at_depth = {}
    child_count = {}
    for i, r in enumerate(rows):
        d = r["depth"]
        p = last_at_depth.get(d - 1)
        if d > 0 and p is not None:
            parent_of[i] = p
            child_count[p] = child_count.get(p, 0) + 1
            pn = rows[p]["n"]
            share[i] = (rows[i]["n"] / pn) if pn else None
        last_at_depth[d] = i
    sole = [parent_of[i] is not None and child_count[parent_of[i]] == 1 for i in range(n)]
    return share, sole


def write_sheet_v2(ws, header_label, rows, base_n, n_responses, n_unassigned):
    """Write one readout to a worksheet — v2 styling. Same contract as
    `view_codebook.write_xlsx_sheet`, so it generalizes to any sheet (nh columns).
    The within-parent `% ouder` column (and inline shares) appear only where the
    sheet nests (nh > 1); a flat sheet like Codeboek (nh == 1) omits them."""
    hier = header_label.split(" / ")                 # ["domain","facet","attribute"] | ["code"]
    nh = len(hier)
    nested = nh > 1
    cols = hier + ["val", "n bruto", "% bruto"] + (["% ouder"] if nested else []) \
        + ["n netto", "% netto", "% (+)", "% (-)"]
    ncol = len(cols)
    col = {name: i for i, name in enumerate(cols, start=1)}   # 1-based index by header name
    pct_cols = [n for n in ("% bruto", "% ouder", "% netto", "% (+)", "% (-)") if n in col]

    shares, sole = _within_parent_shares(rows)

    ws.append(cols)
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.fill, cell.font = _HDR_FILL, _HDR_FONT
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows):
        d = r["depth"]
        share = shares[i]
        ouder = round(share * 100, 1) / 100 if share is not None else None
        # inline '(NN%)' on facet/attribute labels — derived from the same rounded
        # value as the % ouder column so the two never disagree; sole child suppressed
        suffix = f" ({round(ouder * 100)}%)" if (ouder is not None and not sole[i]) else ""
        hcells = [""] * nh
        if d < nh:
            hcells[d] = BULLETS.get(d, "") + r["label"] + suffix
        pos = round(r["pct_pos"], 1) / 100 if r["n"] else None
        neg = round(r["pct_neg"], 1) / 100 if r["n"] else None
        metrics = [r["valence"], r["n"], round(r["pct_bruto"], 1) / 100]
        if nested:
            metrics.append(ouder)
        metrics += [r["n_resp"], round(r["pct_netto"], 1) / 100, pos, neg]
        ws.append(hcells + metrics)
        ri = ws.max_row
        bold = (d == 0)
        for c in range(1, ncol + 1):
            cell = ws.cell(ri, c)
            if c == col["val"] and r["valence"] in _VAL_COLOR:
                cell.font = Font(bold=True, color=_VAL_COLOR[r["valence"]])
                cell.alignment = Alignment(horizontal="center")
            elif bold:
                cell.font = Font(bold=True)
        ws.cell(ri, col["n bruto"]).number_format = "0"
        ws.cell(ri, col["n netto"]).number_format = "0"
        for name in pct_cols:
            ws.cell(ri, col[name]).number_format = "0.0%"
        ws.row_dimensions[ri].outline_level = min(d, 7)

    last_data = ws.max_row
    netto_base = sum(r["n_resp"] for r in rows if r["depth"] == 0)
    total_metrics = ["", base_n, 1.0] + ([None] if nested else []) + [netto_base, 1.0, None, None]
    ws.append(["TOTAAL"] + [""] * (nh - 1) + total_metrics)
    tr = ws.max_row
    for c in range(1, ncol + 1):
        ws.cell(tr, c).font = Font(bold=True)
    ws.cell(tr, col["n bruto"]).number_format = "0"
    ws.cell(tr, col["% bruto"]).number_format = "0.0%"
    ws.cell(tr, col["n netto"]).number_format = "0"
    ws.cell(tr, col["% netto"]).number_format = "0.0%"
    ws.append([])
    ws.append([f"responses: {n_responses}"])
    if n_unassigned:
        ws.append([f"__UNASSIGNED__ (excl. van %-basis): {n_unassigned}"])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{last_data}"
    ws.sheet_properties.outlinePr.summaryBelow = False
    _apply_widths_v2(ws, nh, ncol, last_data)


def _apply_widths_v2(ws, nh, ncol, last_data):
    """v2 column widths: non-terminal hierarchy levels fixed-narrow; everything
    else (terminal hierarchy level + metric columns) auto-fit, clamped 8–55."""
    for c in range(1, ncol + 1):
        letter = get_column_letter(c)
        if c <= nh - 1:                                          # non-terminal hierarchy level
            ws.column_dimensions[letter].width = HIER_WIDTHS[min(c - 1, len(HIER_WIDTHS) - 1)]
        else:                                                   # terminal hierarchy + metrics
            width = max((len(str(ws.cell(r, c).value)) for r in range(1, last_data + 1)
                         if ws.cell(r, c).value is not None), default=8)
            ws.column_dimensions[letter].width = min(max(width + 2, 8), 55)


# =============================================================================
# ENTRY POINT
# =============================================================================
def _append_leeswijzer(ws):
    """Append a small v2 reading guide below the reused production legend."""
    ws.append([])
    ws.append(["Leeswijzer v2"])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    for line in (
        "•  = facet     –  = attribuut",
        "(NN%) achter een facet/attribuut = aandeel binnen de ouder "
        "(facet binnen domein, attribuut binnen facet), o.b.v. bruto — telt op tot 100% per ouder.",
        "Kolom '% ouder' toont ditzelfde aandeel, sorteerbaar.",
    ):
        ws.append([line])


def export_codebook_v2(filename: str = None, var_name: str = None,
                       sample_size=None) -> Path:
    """Build the v2 workbook (Legenda, Codeboek, Taxonomie grof, Taxonomie fijn)
    and write `..._v2.xlsx`.

    Reuses the production data builders and legend from `view_codebook` (read-only)
    and rebinds its module globals so `load_data` sees the right dataset — the same
    pattern the production `export_codebook` uses.
    """
    vc.FILENAME = TEST_DATA.filename if filename is None else filename
    vc.VARIABLE = TEST_DATA.var_name if var_name is None else var_name
    vc.SAMPLE_SIZE = TEST_DATA.sample_size if sample_size is None else sample_size

    responses, codebook, raw_map, metadata, tax = vc.load_data()
    attr_sources = {
        "consolidated": lambda i: i.assigned_attribute,
        "raw": lambda i: raw_map.get(i.idea_id, ""),
    }

    wb = Workbook()
    wb.remove(wb.active)

    # Tab 1 — Legenda: reuse the production legend, then append the v2 reading guide.
    legend_ws = wb.create_sheet(title="Legenda")
    vc.write_legend_sheet(legend_ws, vc.build_legend(codebook, metadata, tax))
    _append_leeswijzer(legend_ws)

    # Tabs 2-4 — Codeboek, Taxonomie (grof), Taxonomie (fijn), in the production order.
    for _title, sheet_name, header, spec, _suffix in vc.VERSIONS:
        if spec[0] == "groups":
            _, group_by, show_attrs, fold = spec
            rows, base_n, n_resp, n_una = vc.build_groups(responses, codebook, group_by, show_attrs, fold)
        else:
            rows, base_n, n_resp, n_una = vc.build_domain_facet_attr(
                responses, fold_tail=True, attr_of=attr_sources[spec[1]])
        write_sheet_v2(wb.create_sheet(title=sheet_name), header, rows, base_n, n_resp, n_una)

    out_dir = vc.codebook_export_dir()
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{vc._codebook_stem(vc.FILENAME, vc.VARIABLE, vc.SAMPLE_SIZE)}_v2.xlsx"
    wb.save(path)
    print(f"XLSX v2 → {path}")
    return path


if __name__ == "__main__":
    export_codebook_v2()
