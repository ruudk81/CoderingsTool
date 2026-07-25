#%%

"""
View codebook / taxonomy: step-6 assignments in two readouts (each printed + saved
to its own CSV):
  1. CODEBOOK   — codes only                       (_codebook)
  2. TAXONOMIE  — domain → facet → attribute        (_taxonomie)
  3. TAXONOMIE  — domain → facet → RAW attribute     (_taxonomie_raw)

Raw attributes are the pre-consolidation (pre-P7/P8) step-4 assignments, read
from the taxonomy cache's `raw_attribute_assignments`.

Per row, two counts:
  - BRUTO = n ideas (mentions) + % of the readout's idea base — a respondent who
    repeats the same code/attribute counts each time.
  - NETTO = unique respondents (deduped per category) + % normalized per depth so
    each level sums to 100%. Netto does not nest through the hierarchy (a respondent
    in a domain via two facets counts once for the domain but once per facet).
Plus a valence balance x% (+) / y% (-) where (+) = positive+neutral, (-) = negative
(idea-level / bruto). The smallest children per parent (together ≤ OVERIG_TAIL_PCT)
fold into one "overig (k …)" row.

The codebook lens excludes the __UNASSIGNED__ sentinel from the % base (reported
separately); the taxonomy lens covers every idea (each idea has a domain/facet).

Usage:
    cd src && python -m pipeline.step_6_codeAssigner.view_codebook
"""

import csv
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

project_root = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(project_root / "src"))

from utils.cacheManager import CacheManager, generate_enhanced_variable_key
from models import CodeAssignedModel
from models import TaxonomyResultsCache
from models import CodingResultsCache, ExtractionMetadata
from pipeline.step_5_codeGenerator.prompts_codeGenerator import ConsolidatedCode
from pipeline.step_7_export.resultsExporter import build_catalog, ResultsExporter

from test_data import TEST_DATA


# =============================================================================
# CONFIGURATION
# =============================================================================

FILENAME = TEST_DATA.filename
VARIABLE = TEST_DATA.var_name
SAMPLE_SIZE = TEST_DATA.sample_size

OVERIG_TAIL_PCT = 0.10        # smallest children summing to ≤ this share of a parent → "overig"
SAVE_CSV = True
SAVE_XLSX = True              # one workbook, one worksheet per readout

_UNASSIGNED = "__UNASSIGNED__"
_NO_ATTR = "(geen attribuut)"
_NO_FACET = "(geen facet)"
_NO_GROUP = "(geen)"
_NEG_VALENCES = {"-", "-1", "neg", "negative"}


# =============================================================================
# DATA LOADING
# =============================================================================

def load_data():
    """Load step-6 response models, the step-5 codebook, and the step-4 raw
    attribute assignments (idea_id → raw attribute, pre-consolidation)."""
    variable_key = generate_enhanced_variable_key(
        selected_variables=[VARIABLE], is_merged=False, sample_size=SAMPLE_SIZE,
    )
    cm = CacheManager()
    results = cm.load_from_cache(FILENAME, "taxonomy_codes", variable_key, CodeAssignedModel)
    if not results:
        raise FileNotFoundError("No taxonomy_codes cache — run step 6 first.")
    codebook = cm.load_metadata_from_cache(FILENAME, "mece_codes", variable_key, CodingResultsCache)
    if not codebook:
        raise FileNotFoundError("No mece_codes cache — run step 5 first.")
    tax = cm.load_metadata_from_cache(FILENAME, "taxonomy", variable_key, TaxonomyResultsCache)
    raw_map: Dict[str, str] = {}
    if tax:
        for dr in tax.partition_results.values():
            raw_map.update(getattr(dr, "raw_attribute_assignments", {}) or {})
    metadata = cm.load_metadata_from_cache(FILENAME, "extracted_ideas", variable_key, ExtractionMetadata)
    return results, codebook, raw_map, metadata, tax


# =============================================================================
# SHARED HELPERS
# =============================================================================

def _vsign(valence: str) -> str:
    v = (valence or "").lower()
    if v.startswith("pos") or v == "+":
        return "+"
    if v.startswith("neg") or v == "-":
        return "-"
    return "~"


def _derived_sign(pct_neg: float) -> str:
    if pct_neg >= 55:
        return "-"
    if pct_neg <= 45:
        return "+"
    return "~"


def _is_neg(valence: str) -> bool:
    return (valence or "").strip().lower() in _NEG_VALENCES


_CATCHALL = {"other", "overig", "rest", _NO_GROUP, _NO_FACET, _NO_ATTR}


def _is_catchall(label: str) -> bool:
    """A catch-all / placeholder group that should sort to the bottom."""
    return (label or "").strip().lower() in {c.lower() for c in _CATCHALL}


class _Cell:
    """Accumulator for one (n, negatives, respondent set) bucket."""
    __slots__ = ("n", "neg", "resp")

    def __init__(self):
        self.n = 0
        self.neg = 0
        self.resp = set()

    def add(self, rid, neg):
        self.n += 1
        self.neg += neg
        self.resp.add(rid)


def _balance(cell: _Cell):
    if not cell.n:
        return 0.0, 0.0
    return 100.0 * (cell.n - cell.neg) / cell.n, 100.0 * cell.neg / cell.n


def _fold_tail(items, parent_total, fold):
    """items: list of (label, _Cell) sorted however. Return (kept, tail) where
    tail is the smallest children summing to ≤ OVERIG_TAIL_PCT of parent_total."""
    if not fold:
        return list(items), []
    threshold = OVERIG_TAIL_PCT * parent_total
    tail, cum = [], 0
    for label, cell in sorted(items, key=lambda kv: kv[1].n):
        if cell.n == 0 or cum + cell.n <= threshold:
            tail.append((label, cell))
            cum += cell.n
        else:
            break
    tail_labels = {lbl for lbl, _ in tail}
    kept = [(lbl, c) for lbl, c in items if lbl not in tail_labels]
    return kept, tail


def _merge_cells(cells):
    m = _Cell()
    for c in cells:
        m.n += c.n
        m.neg += c.neg
        m.resp |= c.resp
    return m


def _row(depth, label, valence, cell, pct_i):
    pos, neg = _balance(cell)
    return {"depth": depth, "label": label, "valence": valence,
            "n": cell.n, "n_resp": len(cell.resp),        # bruto (ideas) / netto (unique respondents)
            "pct_bruto": pct_i(cell.n), "pct_netto": 0.0,  # pct_netto filled by _normalize_netto
            "pct_pos": pos, "pct_neg": neg}


def _normalize_netto(rows):
    """Netto % = a row's unique-respondent count as a share of all rows at the same
    depth, so each depth sums to 100% (the deduped counterpart of bruto %). Netto
    does not nest through the hierarchy — a respondent in a domain via two facets
    counts once for the domain but once per facet — so it is normalized per depth."""
    totals = defaultdict(int)
    for r in rows:
        totals[r["depth"]] += r["n_resp"]
    for r in rows:
        t = totals[r["depth"]]
        r["pct_netto"] = (100.0 * r["n_resp"] / t) if t else 0.0
    return rows


# =============================================================================
# BUILDERS
# =============================================================================

def build_groups(responses, codebook, group_by, show_attrs, fold_tail):
    """Two-level: group (code|domain) → attribute."""
    is_code = (group_by == "code")
    grp: Dict[str, _Cell] = defaultdict(_Cell)
    cell: Dict[str, Dict[str, _Cell]] = defaultdict(lambda: defaultdict(_Cell))
    resp_with_ideas: set = set()
    n_unassigned = 0

    # Join ideas to the codebook via assigned_code_id (K#) — the codebook's
    # CURRENT name wins, so a renamed code keeps its counts. Ideas without an
    # id (pre-id caches) fall back to their stored name.
    id_to_name: Dict[str, str] = {}
    if is_code:
        for c in codebook.raw_codes:
            d = c if isinstance(c, dict) else c.__dict__
            if d.get("code_id"):
                id_to_name[d["code_id"]] = d["code_name"]

    for resp in responses:
        rid = str(resp.respondent_id)
        for idea in (resp.response_ideas or []):
            resp_with_ideas.add(rid)
            if is_code:
                key = (id_to_name.get(getattr(idea, "assigned_code_id", None))
                       or idea.assigned_code or "").strip()
                if not key or key == _UNASSIGNED:
                    n_unassigned += 1
                    continue
            else:
                key = (getattr(idea, "partition_name", "") or idea.domain or "").strip() or _NO_GROUP
            attr = (idea.assigned_attribute or "").strip() or _NO_ATTR
            neg = _is_neg(idea.valence)
            grp[key].add(rid, neg)
            cell[key][attr].add(rid, neg)

    base_n = sum(c.n for c in grp.values())
    n_responses = len(resp_with_ideas)
    pct_i = lambda n: (100.0 * n / base_n) if base_n else 0.0

    # code view: include defined-but-unused codes + source_attributes
    src_attrs: Dict[str, list] = {}
    if is_code:
        cv = {}
        for c in codebook.raw_codes:
            d = c if isinstance(c, dict) else c.__dict__
            cv[d["code_name"]] = _vsign(d.get("valence", ""))
            src_attrs[d["code_name"]] = d.get("source_attributes", []) or []
        for name in cv:
            grp.setdefault(name, _Cell())

    rows = []
    for key in sorted(grp, key=lambda k: (_is_catchall(k), -grp[k].n, k.lower())):
        gc = grp[key]
        valence = cv.get(key, "~") if is_code else _derived_sign(_balance(gc)[1])
        rows.append(_row(0, key, valence, gc, pct_i))
        if not show_attrs:
            continue
        items = {a: c for a, c in cell.get(key, {}).items()}
        for a in src_attrs.get(key, []):
            items.setdefault(a, _Cell())
        kept, tail = _fold_tail(list(items.items()), gc.n, fold_tail)
        for a, c in sorted(kept, key=lambda kv: (_is_catchall(kv[0]), -kv[1].n)):
            rows.append(_row(1, a, "", c, pct_i))
        if len(tail) >= 2:
            rows.append(_row(1, f"overig ({len(tail)} attrs)", "", _merge_cells([c for _, c in tail]), pct_i))
        elif tail:
            a, c = tail[0]
            rows.append(_row(1, a, "", c, pct_i))

    return _normalize_netto(rows), base_n, n_responses, n_unassigned


def build_domain_facet_attr(responses, fold_tail, attr_of):
    """Three-level taxonomy: domain → facet → attribute (pure taxonomy, no codes).

    attr_of(idea) -> the attribute label (consolidated `assigned_attribute`, or the
    raw pre-consolidation attribute via a idea_id lookup).
    """
    dom: Dict[str, _Cell] = defaultdict(_Cell)
    df: Dict[str, Dict[str, _Cell]] = defaultdict(lambda: defaultdict(_Cell))
    dfa: Dict[str, Dict[str, Dict[str, _Cell]]] = \
        defaultdict(lambda: defaultdict(lambda: defaultdict(_Cell)))
    resp_with_ideas: set = set()

    for resp in responses:
        rid = str(resp.respondent_id)
        for idea in (resp.response_ideas or []):
            resp_with_ideas.add(rid)
            d = (getattr(idea, "partition_name", "") or idea.domain or "").strip() or _NO_GROUP
            f = (idea.facet or "").strip() or _NO_FACET
            a = (attr_of(idea) or "").strip() or _NO_ATTR
            neg = _is_neg(idea.valence)
            dom[d].add(rid, neg)
            df[d][f].add(rid, neg)
            dfa[d][f][a].add(rid, neg)

    base_n = sum(c.n for c in dom.values())
    n_responses = len(resp_with_ideas)
    pct_i = lambda n: (100.0 * n / base_n) if base_n else 0.0

    rows = []
    for d in sorted(dom, key=lambda k: (_is_catchall(k), -dom[k].n, k.lower())):
        rows.append(_row(0, d, _derived_sign(_balance(dom[d])[1]), dom[d], pct_i))
        for f in sorted(df[d], key=lambda k: (_is_catchall(k), -df[d][k].n)):
            rows.append(_row(1, f, _derived_sign(_balance(df[d][f])[1]), df[d][f], pct_i))
            # A facet with a single attribute carries no extra info → show the facet only
            if len(dfa[d][f]) == 1:
                continue
            attrs = list(dfa[d][f].items())
            kept, tail = _fold_tail(attrs, df[d][f].n, fold_tail)
            for a, cell in sorted(kept, key=lambda kv: (_is_catchall(kv[0]), -kv[1].n)):
                rows.append(_row(2, a, "", cell, pct_i))
            if len(tail) >= 2:
                rows.append(_row(2, f"overig ({len(tail)} attrs)", "",
                                 _merge_cells([c for _, c in tail]), pct_i))
            elif tail:
                a, cell = tail[0]
                rows.append(_row(2, a, "", cell, pct_i))

    return _normalize_netto(rows), base_n, n_responses, 0


# =============================================================================
# DISPLAY
# =============================================================================

def _bal(r) -> str:
    return f"{r['pct_pos']:.0f}% (+) / {r['pct_neg']:.0f}% (-)" if r["n"] else ""


def print_readout(title, header_label, rows, base_n, n_responses, n_unassigned, compact=False):
    netto_base = sum(r["n_resp"] for r in rows if r["depth"] == 0)
    print(f"\n\n{'=' * 104}")
    print(f"[{title}]  {FILENAME}")
    print(f"{VARIABLE}  |  {base_n} ideas (base)"
          + (f"; {n_unassigned} unassigned" if n_unassigned else "")
          + f"  |  {n_responses} responses")
    print("bruto = ideeën (mentions)  ·  netto = unieke respondenten (ontdubbeld per categorie, % per niveau)")
    print(f"{'=' * 104}")
    print(f"{header_label:50}{'n bruto':>8}{'% bruto':>8}{'n netto':>8}{'% netto':>8}   balans (+/-)")
    print(f"{'-' * 104}")
    for r in rows:
        indent = "    " * r["depth"]
        label = f"{indent}[{r['valence']}] {r['label']}" if r["valence"] else f"{indent}{r['label']}"
        sep = "\n" if (r["depth"] == 0 and not compact) else ""
        tag = "" if r["n"] else "  (unused)"
        print(f"{sep}{label:50}{r['n']:>8}{r['pct_bruto']:>7.1f}%{r['n_resp']:>8}{r['pct_netto']:>7.1f}%   {_bal(r)}{tag}")
    print(f"\n{'-' * 104}")
    print(f"{'TOTAAL':50}{base_n:>8}{100.0:>7.1f}%{netto_base:>8}{100.0:>7.1f}%")
    if n_unassigned:
        print(f"{'__UNASSIGNED__ (excl. van %-basis)':50}{n_unassigned:>8}")


def codebook_export_dir() -> Path:
    return project_root / "exports" / "codebook"


def _codebook_stem(filename: str, var_name: str, sample_size) -> str:
    base = Path(filename).stem.replace(" ", "_")
    size = sample_size if sample_size is not None else "full"   # avoid "_None" for full samples
    return f"codebook_{base}_{var_name}_{size}"


def codebook_xlsx_path(filename: str, var_name: str, sample_size) -> Path:
    """Canonical codebook workbook path — view_codebook AND the app (app_backend) import this,
    so the name/folder can't drift apart."""
    return codebook_export_dir() / f"{_codebook_stem(filename, var_name, sample_size)}.xlsx"


def save_csv(suffix, header_cols, rows, base_n, n_responses, n_unassigned):
    exports_dir = codebook_export_dir()
    exports_dir.mkdir(parents=True, exist_ok=True)
    csv_path = exports_dir / f"{_codebook_stem(FILENAME, VARIABLE, SAMPLE_SIZE)}_{suffix}.csv"
    netto_base = sum(r["n_resp"] for r in rows if r["depth"] == 0)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["depth", "label", "valence", "n_bruto", "pct_bruto",
                    "n_netto", "pct_netto", "pct_pos_neutral", "pct_neg"])
        for r in rows:
            w.writerow([r["depth"], r["label"], r["valence"], r["n"],
                        f"{r['pct_bruto']:.1f}", r["n_resp"], f"{r['pct_netto']:.1f}",
                        f"{r['pct_pos']:.1f}" if r["n"] else "",
                        f"{r['pct_neg']:.1f}" if r["n"] else ""])
        w.writerow(["", "TOTAAL", "", base_n, "100.0", netto_base, "100.0", "", ""])
        w.writerow(["", _UNASSIGNED, "", n_unassigned, "", "", "", "", ""])
        w.writerow(["", "responses", "", n_responses, "", "", "", "", ""])
    print(f"  CSV → {csv_path.name}")


# =============================================================================
# XLSX EXPORT
# =============================================================================

_HDR_FILL = PatternFill("solid", fgColor="366092")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_VAL_COLOR = {"+": "2E7D32", "-": "C62828", "~": "777777"}

# Non-terminal hierarchy levels get a fixed narrow width: only one hierarchy
# column is filled per row, so a non-terminal label (domein, facet, …) overflows
# into the blank cells to its right and stays fully readable, while the tree
# reads as one compact block. The terminal level (attribuut) borders the filled
# metric columns, so it keeps an auto width. Widths are per non-terminal level;
# the list extends by repeating its last value.
HIER_WIDTHS = [12, 18]   # domein, facet

# Per-level bullet prefix on the hierarchy labels, so facets and attributes are
# visually marked and distinct. Keyed by depth (0 = domein stays unmarked).
BULLETS = {1: "• ", 2: "– "}   # facet = gevuld rondje, attribuut = streepje


def _within_parent_shares(rows):
    """Bruto share of each row within its parent: n(row) / n(parent). The parent
    is the nearest preceding row one level up (rows are a pre-order traversal:
    domain → its facets → each facet's attributes). Returns two aligned lists:
    `share` (fraction, or None for depth-0 / no parent) and `sole` (True when the
    parent has exactly one child — used to suppress a trivial inline '(100%)')."""
    n = len(rows)
    share = [None] * n
    parent_of = [None] * n
    last_at_depth = {}
    child_count = {}
    for i, r in enumerate(rows):
        d = r["depth"]
        p = last_at_depth.get(d - 1)
        if d > 0 and p is not None:
            parent_of[i] = p
            child_count[p] = child_count.get(p, 0) + 1
            pn = rows[p]["n"]
            share[i] = (rows[i]["n"] / pn) if pn else None
        last_at_depth[d] = i
    sole = [parent_of[i] is not None and child_count[parent_of[i]] == 1 for i in range(n)]
    return share, sole


def write_xlsx_sheet(ws, header_label, rows, base_n, n_responses, n_unassigned):
    """Write one readout to a worksheet: narrow hierarchy columns with overflow,
    numeric metrics, collapsible row groups, coloured valence. The within-parent
    `% ouder` column (and inline shares) appear only where the sheet nests
    (nh > 1); a flat sheet like Codeboek (nh == 1) omits them."""
    hier = header_label.split(" / ")                 # ["domain","facet","attribute"] | ["code"]
    nh = len(hier)
    nested = nh > 1
    cols = hier + ["val", "n bruto", "% bruto"] + (["% ouder"] if nested else []) \
        + ["n netto", "% netto", "% (+)", "% (-)"]
    ncol = len(cols)
    col = {name: i for i, name in enumerate(cols, start=1)}   # 1-based index by header name
    pct_cols = [n for n in ("% bruto", "% ouder", "% netto", "% (+)", "% (-)") if n in col]

    shares, sole = _within_parent_shares(rows)

    ws.append(cols)
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.fill, cell.font = _HDR_FILL, _HDR_FONT
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows):
        d = r["depth"]
        share = shares[i]
        ouder = round(share * 100, 1) / 100 if share is not None else None
        # inline '(NN%)' on facet/attribute labels — derived from the same rounded
        # value as the % ouder column so the two never disagree; sole child suppressed
        suffix = f" ({round(ouder * 100)}%)" if (ouder is not None and not sole[i]) else ""
        hcells = [""] * nh
        if d < nh:
            hcells[d] = BULLETS.get(d, "") + r["label"] + suffix
        pos = round(r["pct_pos"], 1) / 100 if r["n"] else None
        neg = round(r["pct_neg"], 1) / 100 if r["n"] else None
        metrics = [r["valence"], r["n"], round(r["pct_bruto"], 1) / 100]
        if nested:
            metrics.append(ouder)
        metrics += [r["n_resp"], round(r["pct_netto"], 1) / 100, pos, neg]
        ws.append(hcells + metrics)
        ri = ws.max_row
        bold = (d == 0)
        for c in range(1, ncol + 1):
            cell = ws.cell(ri, c)
            if c == col["val"] and r["valence"] in _VAL_COLOR:
                cell.font = Font(bold=True, color=_VAL_COLOR[r["valence"]])
                cell.alignment = Alignment(horizontal="center")
            elif bold:
                cell.font = Font(bold=True)
        ws.cell(ri, col["n bruto"]).number_format = "0"
        ws.cell(ri, col["n netto"]).number_format = "0"
        for name in pct_cols:
            ws.cell(ri, col[name]).number_format = "0.0%"
        ws.row_dimensions[ri].outline_level = min(d, 7)

    last_data = ws.max_row
    netto_base = sum(r["n_resp"] for r in rows if r["depth"] == 0)
    total_metrics = ["", base_n, 1.0] + ([None] if nested else []) + [netto_base, 1.0, None, None]
    ws.append(["TOTAAL"] + [""] * (nh - 1) + total_metrics)
    tr = ws.max_row
    for c in range(1, ncol + 1):
        ws.cell(tr, c).font = Font(bold=True)
    ws.cell(tr, col["n bruto"]).number_format = "0"
    ws.cell(tr, col["% bruto"]).number_format = "0.0%"
    ws.cell(tr, col["n netto"]).number_format = "0"
    ws.cell(tr, col["% netto"]).number_format = "0.0%"
    ws.append([])
    ws.append([f"responses: {n_responses}"])
    if n_unassigned:
        ws.append([f"__UNASSIGNED__ (excl. van %-basis): {n_unassigned}"])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{last_data}"
    ws.sheet_properties.outlinePr.summaryBelow = False
    _apply_widths(ws, nh, ncol, last_data)


def _apply_widths(ws, nh, ncol, last_data):
    """Column widths: non-terminal hierarchy levels fixed-narrow; everything
    else (terminal hierarchy level + metric columns) auto-fit, clamped 8–55."""
    for c in range(1, ncol + 1):
        letter = get_column_letter(c)
        if c <= nh - 1:                                          # non-terminal hierarchy level
            ws.column_dimensions[letter].width = HIER_WIDTHS[min(c - 1, len(HIER_WIDTHS) - 1)]
        else:                                                   # terminal hierarchy + metrics
            width = max((len(str(ws.cell(r, c).value)) for r in range(1, last_data + 1)
                         if ws.cell(r, c).value is not None), default=8)
            ws.column_dimensions[letter].width = min(max(width + 2, 8), 55)


def save_xlsx(wb):
    exports_dir = codebook_export_dir()
    exports_dir.mkdir(parents=True, exist_ok=True)
    path = codebook_xlsx_path(FILENAME, VARIABLE, SAMPLE_SIZE)
    wb.save(path)
    print(f"\nXLSX → {path}")
    return path


def _append_leeswijzer(ws):
    """Append a small reading guide below the legend."""
    ws.append([])
    ws.append(["Leeswijzer"])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    for line in (
        "•  = facet     –  = attribuut",
        "(NN%) achter een facet/attribuut = aandeel binnen de ouder "
        "(facet binnen domein, attribuut binnen facet), o.b.v. bruto — telt op tot 100% per ouder.",
        "Kolom '% ouder' toont ditzelfde aandeel, sorteerbaar.",
    ):
        ws.append([line])


# =============================================================================
# MAIN
# =============================================================================

# (title, sheet_name, header_label, builder spec, csv_suffix)
VERSIONS = [
    ("CODEBOOK",             "Codeboek",         "code",                          ("groups", "code", False, False), "codebook"),
    ("TAXONOMIE",             "Taxonomie (grof)", "domain / facet / attribute",     ("dfa", "consolidated"), "taxonomie"),
    ("TAXONOMIE (ruwe attr)", "Taxonomie (fijn)", "domain / facet / raw attribute", ("dfa", "raw"),          "taxonomie_raw"),
]

def export_codebook(filename: str = None, var_name: str = None,
                    sample_size: Optional[int] = None, *,
                    write_csv: bool = SAVE_CSV, write_xlsx: bool = SAVE_XLSX,
                    print_console: bool = False) -> Optional[Path]:
    """Write the codebook/taxonomy readouts (CSV + XLSX) to exports/codebook/.

    Runs AFTER step 6 — reads taxonomy_codes (6) + mece_codes (5) + taxonomy (4) +
    extracted_ideas (3) from cache. Returns the xlsx path (None if write_xlsx=False).

    Dataset params default to TEST_DATA (so the standalone `python -m ...view_codebook`
    dev run is unchanged); the app passes them explicitly. Rebinds the module globals
    once so load_data/save_csv/save_xlsx (which read them in-body) see the right dataset.
    """
    global FILENAME, VARIABLE, SAMPLE_SIZE
    FILENAME = FILENAME if filename is None else filename
    VARIABLE = VARIABLE if var_name is None else var_name
    SAMPLE_SIZE = SAMPLE_SIZE if sample_size is None else sample_size

    responses, codebook, raw_map, metadata, tax = load_data()
    attr_sources = {
        "consolidated": lambda i: i.assigned_attribute,
        "raw": lambda i: raw_map.get(i.idea_id, ""),
    }
    wb = Workbook() if write_xlsx else None
    if wb is not None:
        wb.remove(wb.active)
        # Legenda: step 7's canonical catalog legend (Codeboek + taxonomy A-E, with
        # definitions) — build_catalog is the single source of truth, so the legend
        # always matches the tabs. A reading guide is appended below it.
        codes = [ConsolidatedCode(**c) if isinstance(c, dict) else c for c in (codebook.raw_codes or [])]
        cat = build_catalog(codes, codebook.partition_set, codebook.partition_results,
                            metadata, responses, tax)
        legend_ws = wb.create_sheet(title="Legenda")
        ResultsExporter(verbose=False)._write_legend_sheet(legend_ws, cat)
        _append_leeswijzer(legend_ws)
    for title, sheet_name, header, spec, suffix in VERSIONS:
        if spec[0] == "groups":
            _, group_by, show_attrs, fold = spec
            rows, base_n, n_resp, n_una = build_groups(
                responses, codebook, group_by, show_attrs, fold)
            compact = not show_attrs
        else:
            rows, base_n, n_resp, n_una = build_domain_facet_attr(
                responses, fold_tail=True, attr_of=attr_sources[spec[1]])
            compact = False
        if print_console:
            print_readout(title, header, rows, base_n, n_resp, n_una, compact)
        if write_csv:
            save_csv(suffix, header, rows, base_n, n_resp, n_una)
        if wb is not None:
            write_xlsx_sheet(wb.create_sheet(title=sheet_name), header,
                             rows, base_n, n_resp, n_una)
    return save_xlsx(wb) if wb is not None else None


if __name__ == "__main__":
    # Standalone dev run: uses TEST_DATA + prints the readouts to the terminal.
    export_codebook(print_console=True)

# %%
