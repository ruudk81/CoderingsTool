"""
Step 7 — Export.

Three analytical outputs, each as an Excel worksheet AND a .sav file, plus an
Excel legend sheet:
  - Output 1 "codeboek"     — per respondent, dichotomous code matrix
  - Output 2 "taxonomie"    — per respondent, dichotomous domain/facet/attribute matrix
  - Output 3 "gecombineerd" — per expressed idea, long (the source for 1 & 2)

Design + decisions: see dev/WORK_TO_BE_DONE.md.
"""

import math
import re
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Any, List, Dict, Optional, Tuple
from pathlib import Path

import numpy as np
import pandas as pd
import pyreadstat
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from pipeline.step_6_codeAssigner.models_codeAssigner import CodeAssignedModel
from pipeline.step_5_codeGenerator.prompts_codeGenerator import ConsolidatedCode
from pipeline.step_4_classifier.models_classifier import DomainSet, DomainResultModel
from utils.verboseReporter import VerboseReporter


# === CONSTANTS ====================================================================================================
UNASSIGNED_SENTINEL = "__UNASSIGNED__"

# Quality-filter codes (8-digit) from step 2 — treated as "codes" in output 1/3.
FILTER_CODE_LABELS = {
    99999997: "Weet niet / geen antwoord",
    99999998: "Geen antwoord / leeg",
    99999999: "Betekenisloos",
}

DICHOTOMOUS_VALUE_LABELS = {0: "Niet Genoemd", 1: "Wel Genoemd"}
VALENCE_VALUE_LABELS = {-1: "negatief", 0: "neutraal", 1: "positief"}

# Excel styling (matches step_6/view_codebook.py)
_BLOCK_FILL = PatternFill("solid", fgColor="366092")   # dark blue
_BLOCK_FONT = Font(bold=True, color="FFFFFF", size=12)
_SUB_FILL = PatternFill("solid", fgColor="8EAADB")     # lighter blue
_SUB_FONT = Font(bold=True, color="1F3864")
_COLH_FILL = PatternFill("solid", fgColor="D9E1F2")    # very light blue
_COLH_FONT = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor="366092")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_WRAP = Alignment(wrap_text=True, vertical="top")


# === OUTPUT PATHS (single source of truth — the app imports these) ================================
def results_export_dir(export_dir=None) -> Path:
    """Folder the results deliverables go in (default exports/coderingen/)."""
    if export_dir is None:
        export_dir = Path(__file__).resolve().parents[3] / "exports" / "coderingen"
    return Path(export_dir)


def results_xlsx_path(filename: str, var_name: str, export_dir=None) -> Path:
    """Canonical path of the results workbook. ResultsExporter.export() AND the app
    (app_backend.export_path) call THIS, so the name/folder can't drift apart."""
    base = f"{Path(filename).stem}_{var_name}"
    return results_export_dir(export_dir) / f"{base}_codering.xlsx"


def _green_scale() -> ColorScaleRule:
    """0 -> light green, 1 -> dark green (fresh instance per use)."""
    return ColorScaleRule(start_type="num", start_value=0, start_color="C6EFCE",
                          end_type="num", end_value=1, end_color="2E7D32")


def _autofit(ws, col_idx: int, cap: int, pad: int = 2, max_row: Optional[int] = None):
    """Set a column's width to its longest cell value (capped)."""
    longest = 0
    for r in range(1, (max_row or ws.max_row) + 1):
        v = ws.cell(r, col_idx).value
        if v is not None:
            longest = max(longest, len(str(v)))
    ws.column_dimensions[get_column_letter(col_idx)].width = min(longest + pad, cap)


def _valence_to_num(v: Optional[str]) -> Optional[int]:
    """'+' -> 1, '-' -> -1, '0'/'' -> 0, None -> None."""
    if v is None:
        return None
    s = str(v).strip()
    if s == "+":
        return 1
    if s == "-":
        return -1
    return 0


def _clean_var_lab(lab: Optional[str]) -> str:
    """Strip a leading empty-bracket artifact (e.g. '[]Q1t ...' -> 'Q1t ...').

    No-op when the label has no leading empty brackets; real content like '[Q1]'
    is preserved. Safe on None/empty."""
    if not lab:
        return ""
    return re.sub(r"^\s*\[\s*\]\s*", "", lab).strip()


def _clean_response(text: Any) -> str:
    """Render a raw response for the 'original variable' column; NA-likes -> ''."""
    if text is None:
        return ""
    s = str(text).strip()
    if s.lower() in ("nan", "<na>", "none"):
        return ""
    return s


def _sanitize_excel_text(text):
    if not isinstance(text, str):
        return text
    return "".join(c if ord(c) > 31 or c in "\t\n\r" else " " for c in text).replace("\x08", "")


# === CATALOG (canonical numbering, shared across outputs) ==========================================================
@dataclass
class Entry:
    number: int
    name: str
    definition: str = ""
    domain_name: Optional[str] = None
    domain_number: Optional[int] = None


@dataclass
class Catalog:
    codes: Dict[str, Entry] = field(default_factory=dict)            # code_name -> Entry
    domains: Dict[str, Entry] = field(default_factory=dict)          # domain_name -> Entry
    facets: Dict[Tuple[str, str], Entry] = field(default_factory=dict)      # (domain, facet) -> Entry
    attributes: Dict[Tuple[str, str], Entry] = field(default_factory=dict)      # (domain, consolidated attr) -> Entry  [grof]
    attributes_raw: Dict[Tuple[str, str], Entry] = field(default_factory=dict)  # (domain, raw attr) -> Entry          [fijn]
    raw_map: Dict[str, str] = field(default_factory=dict)            # idea_id -> raw attribute name
    dimension: Tuple[str, str] = ("", "")                            # (name, definition)


def build_catalog(
    codes: List[ConsolidatedCode],
    partition_set: DomainSet,
    partition_results: Dict[str, DomainResultModel],
    metadata: Any,
    responses: List[CodeAssignedModel],
    tax: Any = None,
) -> Catalog:
    """Assign canonical numbers to codes / domains / facets / attributes from the cache.

    Facets and attributes are numbered grouped by domain order; each carries its
    domain number. The union with idea-assigned (domain, facet/attribute) ensures
    every idea maps to a catalog entry.

    Two attribute sets:
      - attributes      = consolidated (grof), from step-5 partition_results.attributes
      - attributes_raw  = raw (fijn), from step-4 `tax`.raw_attributes + raw_attribute_assignments
    """
    cat = Catalog()

    # --- Codes (step 5 codebook order) ---
    for i, c in enumerate(codes, 1):
        cat.codes[c.code_name] = Entry(number=i, name=c.code_name, definition=c.definition or "")

    # --- Domains (partition_set order; definitions from inclusion_definition) ---
    domain_def = {p.partition_name: (p.inclusion_definition or "") for p in partition_set.partitions}
    domain_order = [p.partition_name for p in partition_set.partitions]
    for dn in partition_results.keys():
        if dn not in domain_order:
            domain_order.append(dn)
    for i, dn in enumerate(domain_order, 1):
        cat.domains[dn] = Entry(number=i, name=dn, definition=domain_def.get(dn, ""))

    # --- Facets + attributes (grouped by domain order) ---
    fnum = anum = 0
    for dn in domain_order:
        dnum = cat.domains[dn].number
        dr = partition_results.get(dn)
        if not dr:
            continue
        for f in (dr.facets or []):
            if not isinstance(f, dict):
                continue
            fname = f.get("facet_name")
            if fname and (dn, fname) not in cat.facets:
                fnum += 1
                cat.facets[(dn, fname)] = Entry(fnum, fname, f.get("facet_description", ""), dn, dnum)
        for attr_list in (dr.attributes or {}).values():
            for a in (attr_list or []):
                if not isinstance(a, dict):
                    continue
                aname = a.get("attribute_name")
                if aname and (dn, aname) not in cat.attributes:
                    anum += 1
                    cat.attributes[(dn, aname)] = Entry(anum, aname, a.get("attribute_description", ""), dn, dnum)

    # --- Union: add (domain, facet/attribute) seen on ideas but missing from taxonomy ---
    for resp in responses:
        for idea in (resp.response_ideas or []):
            dn = idea.partition_name or idea.domain or ""
            if dn and dn not in cat.domains:
                num = len(cat.domains) + 1
                cat.domains[dn] = Entry(num, dn, "")
            dnum = cat.domains.get(dn).number if dn in cat.domains else None
            if idea.facet and dn and (dn, idea.facet) not in cat.facets:
                fnum += 1
                cat.facets[(dn, idea.facet)] = Entry(fnum, idea.facet, "", dn, dnum)
            if idea.attribute and dn and (dn, idea.attribute) not in cat.attributes:
                anum += 1
                cat.attributes[(dn, idea.attribute)] = Entry(anum, idea.attribute, "", dn, dnum)

    # --- Raw (fijn) attributes: from step-4 tax (raw_attributes desc + raw_attribute_assignments) ---
    raw_desc: Dict[str, str] = {}
    if tax is not None:
        for dr in tax.partition_results.values():
            for attr_list in (getattr(dr, "raw_attributes", {}) or {}).values():
                for a in (attr_list or []):
                    if isinstance(a, dict) and a.get("attribute_name"):
                        raw_desc.setdefault(a["attribute_name"], a.get("attribute_description", ""))
            cat.raw_map.update(getattr(dr, "raw_attribute_assignments", {}) or {})
    # number raw attributes grouped by domain order, sourced from idea (domain, raw attr)
    raw_by_domain: Dict[str, set] = defaultdict(set)
    for resp in responses:
        for idea in (resp.response_ideas or []):
            dn = idea.partition_name or idea.domain or ""
            ra = cat.raw_map.get(idea.idea_id)
            if dn and ra:
                raw_by_domain[dn].add(ra)
    rnum = 0
    for dn_entry in sorted(cat.domains.values(), key=lambda x: x.number):
        for ra in sorted(raw_by_domain.get(dn_entry.name, [])):
            if (dn_entry.name, ra) not in cat.attributes_raw:
                rnum += 1
                cat.attributes_raw[(dn_entry.name, ra)] = Entry(rnum, ra, raw_desc.get(ra, ""),
                                                                dn_entry.name, dn_entry.number)

    if metadata is not None:
        cat.dimension = (getattr(metadata, "primary_dimension", "") or "",
                         getattr(metadata, "primary_dimension_description", "") or "")
    return cat


# === EXPORTER =====================================================================================================
class ResultsExporter:

    def __init__(self, verbose: bool = True):
        self.verbose = verbose
        self.reporter = VerboseReporter(verbose, capture_logging=True)

    # ---- public API ----
    def export(self,
               responses: List[CodeAssignedModel],
               codes: List[ConsolidatedCode],
               partition_set: DomainSet,
               partition_results: Dict[str, DomainResultModel],
               metadata: Any,
               tax: Any = None,
               quality_filtered: Optional[List] = None,
               filename: str = "export",
               var_name: str = "VAR",
               var_lab: str = "",
               export_dir: Optional[str] = None) -> Dict[str, str]:
        """Build the catalog + 3 outputs, write one Excel workbook (4 sheets) and 3 .sav files."""
        self.reporter.section_header("EXPORT — codeboek / taxonomie / gecombineerd")
        var_lab = _clean_var_lab(var_lab)

        cat = build_catalog(codes, partition_set, partition_results, metadata, responses, tax)
        self.reporter.stat_line(
            f"Catalog: {len(cat.codes)} codes, {len(cat.domains)} domains, {len(cat.facets)} facets, "
            f"{len(cat.attributes)} attributes (grof) / {len(cat.attributes_raw)} (fijn)")

        filtered = [r for r in (quality_filtered or []) if getattr(r, "quality_filter", False)]
        resp_text = {r.respondent_id: _clean_response(r.response) for r in (quality_filtered or [])}

        long_df, long_vlabels, long_collabels = self._build_long(responses, filtered, cat, var_name, var_lab, resp_text)
        codes_df, codes_vlabels, codes_collabels = self._build_codes_matrix(responses, filtered, cat, var_name, var_lab, resp_text)
        grof_df, grof_vlabels, grof_collabels = self._build_taxonomy_matrix(
            responses, filtered, cat, var_name, var_lab, resp_text,
            cat.attributes, lambda idea: idea.attribute)
        fijn_df, fijn_vlabels, fijn_collabels = self._build_taxonomy_matrix(
            responses, filtered, cat, var_name, var_lab, resp_text,
            cat.attributes_raw, lambda idea: cat.raw_map.get(idea.idea_id))

        # output paths — final deliverables go in their own subfolder
        export_dir = results_export_dir(export_dir)
        export_dir.mkdir(parents=True, exist_ok=True)
        base = f"{Path(filename).stem}_{var_name}"

        # Excel (one workbook, 5 sheets)
        xlsx_path = results_xlsx_path(filename, var_name, export_dir)
        self._write_excel(xlsx_path, cat, codes_df, codes_collabels,
                          grof_df, grof_collabels, fijn_df, fijn_collabels, long_df)

        # 4 .sav files
        paths = {"excel": str(xlsx_path)}
        for suffix, df, vlab, clab in [
            ("codeboek", codes_df, codes_vlabels, codes_collabels),
            ("taxonomie_grof", grof_df, grof_vlabels, grof_collabels),
            ("taxonomie_fijn", fijn_df, fijn_vlabels, fijn_collabels),
            ("gecombineerd", long_df, long_vlabels, long_collabels),
        ]:
            sav_path = export_dir / f"{base}_{suffix}.sav"
            self._write_sav(df, sav_path, vlab, clab)
            paths[suffix] = str(sav_path)

        self.reporter.stat_line(f"Excel: {xlsx_path.name}  ({len(long_df)} idea-rows, {len(codes_df)} respondents)")
        for s in ("codeboek", "taxonomie_grof", "taxonomie_fijn", "gecombineerd"):
            self.reporter.stat_line(f".sav: {Path(paths[s]).name}")
        return paths

    # ---- output 3: long (per idea) — the source ----
    def _build_long(self, responses, filtered, cat: Catalog, var_name, var_lab, resp_text):
        rows = []
        for resp in responses:
            for idea in (resp.response_ideas or []):
                dn = idea.partition_name or idea.domain or ""
                code_entry = cat.codes.get(idea.assigned_code) if idea.assigned_code and idea.assigned_code != UNASSIGNED_SENTINEL else None
                dom = cat.domains.get(dn)
                fac = cat.facets.get((dn, idea.facet)) if idea.facet else None
                att = cat.attributes.get((dn, idea.attribute)) if idea.attribute else None
                rows.append({
                    "DLNMID": resp.respondent_id,
                    var_name: _clean_response(resp.response),
                    "code": code_entry.number if code_entry else np.nan,
                    "domain": dom.number if dom else np.nan,
                    "facet": fac.number if fac else np.nan,
                    "attribute": att.number if att else np.nan,
                    "valence": _valence_to_num(idea.valence) if idea.valence else np.nan,
                    "instance": _sanitize_excel_text(idea.instance or ""),
                    "interpretation": _sanitize_excel_text(idea.interpretation or ""),
                    "abstraction": _sanitize_excel_text(idea.abstraction or ""),
                })
        # filtered respondents → one row, code = filter code, rest N/A
        for r in filtered:
            rows.append({
                "DLNMID": r.respondent_id,
                var_name: resp_text.get(r.respondent_id, ""),
                "code": r.quality_filter_code,
                "domain": np.nan, "facet": np.nan, "attribute": np.nan, "valence": np.nan,
                "instance": "", "interpretation": "", "abstraction": "",
            })
        df = pd.DataFrame(rows, columns=[
            "DLNMID", var_name, "code", "domain", "facet", "attribute",
            "valence", "instance", "interpretation", "abstraction"])

        code_vl = {e.number: e.name for e in cat.codes.values()}
        code_vl.update(FILTER_CODE_LABELS)
        value_labels = {
            "code": code_vl,
            "domain": {e.number: e.name for e in cat.domains.values()},
            "facet": {e.number: e.name for e in cat.facets.values()},          # bare names in output 3
            "attribute": {e.number: e.name for e in cat.attributes.values()},  # bare names in output 3
            "valence": VALENCE_VALUE_LABELS,
        }
        col_labels = {"DLNMID": "Respondent ID", var_name: var_lab or var_name,
                      "code": "Code", "domain": "Domein", "facet": "Facet",
                      "attribute": "Attribuut", "valence": "Valentie",
                      "instance": "Instance", "interpretation": "Interpretatie",
                      "abstraction": "Abstractie"}
        return df, value_labels, col_labels

    # ---- output 1: codes per respondent (dichotomous) ----
    def _build_codes_matrix(self, responses, filtered, cat: Catalog, var_name, var_lab, resp_text):
        ideas_by_resp = {r.respondent_id: (r.response_ideas or []) for r in responses}
        resp_response = {r.respondent_id: _clean_response(r.response) for r in responses}

        code_cols = [(f"{var_name}code_{e.number}", e.name, e.number) for e in sorted(cat.codes.values(), key=lambda x: x.number)]
        filter_cols = [(f"{var_name}code_{fc}", lbl, fc) for fc, lbl in FILTER_CODE_LABELS.items()]

        rows = []
        for r in (responses_and_filtered := self._respondent_order(responses, filtered)):
            rid, is_filtered, fcode = r
            row = {"DLNMID": rid, var_name: resp_text.get(rid) or resp_response.get(rid, "")}
            if is_filtered:
                for col, _, _ in code_cols:
                    row[col] = np.nan                      # real codes N/A for filtered
                for col, _, fc in filter_cols:
                    row[col] = 1 if fc == fcode else 0
            else:
                assigned = {i.assigned_code for i in ideas_by_resp.get(rid, []) if i.assigned_code and i.assigned_code != UNASSIGNED_SENTINEL}
                assigned_nums = {cat.codes[c].number for c in assigned if c in cat.codes}
                for col, _, num in code_cols:
                    row[col] = 1 if num in assigned_nums else 0
                for col, _, _ in filter_cols:
                    row[col] = 0
            rows.append(row)

        ordered_cols = ["DLNMID", var_name] + [c for c, _, _ in code_cols] + [c for c, _, _ in filter_cols]
        df = pd.DataFrame(rows, columns=ordered_cols)
        col_labels = {"DLNMID": "Respondent ID", var_name: var_lab or var_name}
        col_labels.update({c: lbl for c, lbl, _ in code_cols})
        col_labels.update({c: lbl for c, lbl, _ in filter_cols})
        value_labels = {c: DICHOTOMOUS_VALUE_LABELS for c, _, _ in (code_cols + filter_cols)}
        return df, value_labels, col_labels

    # ---- output 2: taxonomy per respondent (dichotomous) — grof or fijn ----
    def _build_taxonomy_matrix(self, responses, filtered, cat: Catalog, var_name, var_lab, resp_text,
                               attr_catalog: Dict[Tuple[str, str], Entry], attr_getter):
        """Domain + facet (consolidated) + attribute. `attr_catalog`/`attr_getter`
        select consolidated (grof) or raw (fijn) attributes; domain/facet are shared."""
        ideas_by_resp = {r.respondent_id: (r.response_ideas or []) for r in responses}
        resp_response = {r.respondent_id: _clean_response(r.response) for r in responses}

        dom_cols = [(f"{var_name}domain_{e.number}", e.name, ("domain", e.number))
                    for e in sorted(cat.domains.values(), key=lambda x: x.number)]
        fac_cols = [(f"{var_name}facet_{e.number}_{e.domain_number}", f"{e.name}_{e.domain_number}", ("facet", e.number))
                    for e in sorted(cat.facets.values(), key=lambda x: x.number)]
        att_cols = [(f"{var_name}attr_{e.number}_{e.domain_number}", f"{e.name}_{e.domain_number}", ("attribute", e.number))
                    for e in sorted(attr_catalog.values(), key=lambda x: x.number)]
        all_cols = dom_cols + fac_cols + att_cols

        rows = []
        for rid, is_filtered, _ in self._respondent_order(responses, filtered):
            row = {"DLNMID": rid, var_name: resp_text.get(rid) or resp_response.get(rid, "")}
            if is_filtered:
                for col, _, _ in all_cols:
                    row[col] = np.nan
            else:
                hit = {"domain": set(), "facet": set(), "attribute": set()}
                for idea in ideas_by_resp.get(rid, []):
                    dn = idea.partition_name or idea.domain or ""
                    if dn in cat.domains:
                        hit["domain"].add(cat.domains[dn].number)
                    if idea.facet and (dn, idea.facet) in cat.facets:
                        hit["facet"].add(cat.facets[(dn, idea.facet)].number)
                    aname = attr_getter(idea)
                    if aname and (dn, aname) in attr_catalog:
                        hit["attribute"].add(attr_catalog[(dn, aname)].number)
                for col, _, (kind, num) in all_cols:
                    row[col] = 1 if num in hit[kind] else 0
            rows.append(row)

        ordered_cols = ["DLNMID", var_name] + [c for c, _, _ in all_cols]
        df = pd.DataFrame(rows, columns=ordered_cols)
        col_labels = {"DLNMID": "Respondent ID", var_name: var_lab or var_name}
        col_labels.update({c: lbl for c, lbl, _ in all_cols})
        value_labels = {c: DICHOTOMOUS_VALUE_LABELS for c, _, _ in all_cols}
        return df, value_labels, col_labels

    @staticmethod
    def _respondent_order(responses, filtered):
        """All respondents: valid (from step 6) then filtered (from step 2).

        Returns list of (respondent_id, is_filtered, filter_code)."""
        out = [(r.respondent_id, False, None) for r in responses]
        out += [(r.respondent_id, True, r.quality_filter_code) for r in filtered]
        return out

    # ---- Excel ----
    def _write_excel(self, path, cat: Catalog, codes_df, codes_labels,
                     grof_df, grof_labels, fijn_df, fijn_labels, long_df):
        wb = Workbook()
        self._write_legend_sheet(wb.active, cat)
        wb.active.title = "Legenda"
        # dichotomous sheets: use readable labels as column headers (var names are cryptic)
        for name, df, labels in [
            ("Codeboek", codes_df, codes_labels),
            ("Taxonomie grof", grof_df, grof_labels),
            ("Taxonomie fijn", fijn_df, fijn_labels),
        ]:
            ws = wb.create_sheet(name)
            self._write_data_sheet(ws, df, cat, kind="matrix", header_labels=labels)
            self._style_matrix_sheet(ws, df, labels)
        # long sheet: column names are already readable
        ws_long = wb.create_sheet("Gecombineerd")
        self._write_data_sheet(ws_long, long_df, cat, kind="long")
        self._style_long_sheet(ws_long, long_df)
        wb.save(path)

    def _style_matrix_sheet(self, ws, df, header_labels):
        """Output 1/2: col A = DLNMID, col B fits its variable label, dichotomous
        columns (C..) get a light->dark green color scale (1 = green)."""
        ws.column_dimensions["A"].width = 12
        header_b = (header_labels or {}).get(df.columns[1], df.columns[1])
        ws.column_dimensions["B"].width = min(len(str(header_b)) + 2, 80)
        n = len(df.columns)
        if n >= 3:
            rng = f"C2:{get_column_letter(n)}{len(df) + 1}"
            ws.conditional_formatting.add(rng, _green_scale())

    def _style_long_sheet(self, ws, df):
        """Output 3: B = response (75), C-G coded fields autofit, H-J free text (50)."""
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 75
        for ci in range(3, 8):       # C-G: code, domain, facet, attribute, valence
            _autofit(ws, ci, cap=70)
        for ci in range(8, 11):      # H-J: instance, interpretation, abstraction
            ws.column_dimensions[get_column_letter(ci)].width = 50

    def _write_legend_sheet(self, ws, cat: Catalog):
        NCOL = 3
        r = 1

        def banner(text, fill, font):
            nonlocal r
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
            c = ws.cell(r, 1, text); c.fill = fill; c.font = font
            r += 1

        def colheader(name):
            nonlocal r
            for i, h in enumerate(["nr", name, "definitie"], 1):
                c = ws.cell(r, i, h); c.fill = _COLH_FILL; c.font = _COLH_FONT
            r += 1

        def items(entries):
            nonlocal r
            for e in entries:
                ws.cell(r, 1, e.number)
                ws.cell(r, 2, e.name)
                c = ws.cell(r, 3, e.definition); c.alignment = _WRAP
                r += 1

        # CODEBOEK
        banner("CODEBOEK", _BLOCK_FILL, _BLOCK_FONT)
        colheader("code")
        items(sorted(cat.codes.values(), key=lambda x: x.number))
        for fc, lbl in FILTER_CODE_LABELS.items():
            ws.cell(r, 1, fc); ws.cell(r, 2, lbl)
            ws.cell(r, 3, "Filtercategorie uit kwaliteitsfilter (stap 2)").alignment = _WRAP
            r += 1
        r += 1  # blank

        # TAXONOMIE
        banner("TAXONOMIE", _BLOCK_FILL, _BLOCK_FONT)
        banner("A — Dimensie", _SUB_FILL, _SUB_FONT)
        colheader("dimensie")
        ws.cell(r, 1, 1); ws.cell(r, 2, cat.dimension[0])
        ws.cell(r, 3, cat.dimension[1]).alignment = _WRAP
        r += 2
        banner("B — Domeinen", _SUB_FILL, _SUB_FONT)
        colheader("domein")
        items(sorted(cat.domains.values(), key=lambda x: x.number))
        r += 1
        # C — Facetten (grouped by domain)
        banner("C — Facetten", _SUB_FILL, _SUB_FONT)
        for dn in sorted(cat.domains.values(), key=lambda x: x.number):
            facs = [e for e in cat.facets.values() if e.domain_name == dn.name]
            if not facs:
                continue
            banner(dn.name, _COLH_FILL, _COLH_FONT)
            colheader("facet")
            items(sorted(facs, key=lambda x: x.number))
        r += 1
        # D — Attributen grof (consolidated, grouped by domain)
        banner("D — Attributen (grof / geconsolideerd)", _SUB_FILL, _SUB_FONT)
        for dn in sorted(cat.domains.values(), key=lambda x: x.number):
            atts = [e for e in cat.attributes.values() if e.domain_name == dn.name]
            if not atts:
                continue
            banner(dn.name, _COLH_FILL, _COLH_FONT)
            colheader("attribuut")
            items(sorted(atts, key=lambda x: x.number))
        r += 1
        # E — Attributen fijn (raw, grouped by domain)
        banner("E — Attributen (fijn / ruw)", _SUB_FILL, _SUB_FONT)
        for dn in sorted(cat.domains.values(), key=lambda x: x.number):
            atts = [e for e in cat.attributes_raw.values() if e.domain_name == dn.name]
            if not atts:
                continue
            banner(dn.name, _COLH_FILL, _COLH_FONT)
            colheader("attribuut")
            items(sorted(atts, key=lambda x: x.number))

        ws.column_dimensions["A"].width = 6
        _autofit(ws, 2, cap=255)   # labels — fit to longest
        _autofit(ws, 3, cap=255)   # definitions — fit to longest (255 = Excel's hard max)

    def _write_data_sheet(self, ws, df, cat: Catalog, kind: str, header_labels: Optional[Dict] = None):
        """Write a data sheet. For output 3 (long), numeric coded columns are shown
        as readable labels; dichotomous matrices keep 0/1 (blank for N/A)."""
        label_maps = {}
        if kind == "long":
            code_vl = {e.number: e.name for e in cat.codes.values()}; code_vl.update(FILTER_CODE_LABELS)
            label_maps = {
                "code": code_vl,
                "domain": {e.number: e.name for e in cat.domains.values()},
                "facet": {e.number: e.name for e in cat.facets.values()},
                "attribute": {e.number: e.name for e in cat.attributes.values()},
                "valence": VALENCE_VALUE_LABELS,
            }
        cols = list(df.columns)
        for j, col in enumerate(cols, 1):
            h = header_labels.get(col, col) if header_labels else col
            c = ws.cell(1, j, h); c.fill = _HDR_FILL; c.font = _HDR_FONT
        for i, (_, row) in enumerate(df.iterrows(), 2):
            for j, col in enumerate(cols, 1):
                val = row[col]
                if isinstance(val, float) and math.isnan(val):
                    continue  # blank for N/A
                if col in label_maps and val == val:
                    val = label_maps[col].get(int(val), val)
                ws.cell(i, j, _sanitize_excel_text(val) if isinstance(val, str) else val)
        ws.freeze_panes = "A2"

    # ---- .sav ----
    def _write_sav(self, df, path, value_labels, col_labels):
        out = df.copy()
        # pyreadstat wants column_labels aligned to df columns
        column_labels = [col_labels.get(c, c) for c in out.columns]
        pyreadstat.write_sav(
            out, str(path),
            column_labels=column_labels,
            variable_value_labels={k: v for k, v in value_labels.items() if k in out.columns},
        )
